#!/usr/bin/env python3
"""
REVISION ANALYSES — PREDIX HER2 MULTIMODAL pCR PREDICTION
=========================================================
Statistical analyses added during peer review. Every number this module
produces is derived from the PKL artefacts written by
`multimodal_pcr_pipeline.py`; nothing is recomputed from a model fit, so the
outputs cannot drift from the pipeline that generated them.

WHY THIS MODULE EXISTS
----------------------
Four specific criticisms required analyses that the original code did not
perform. Each has a section below and a corresponding output artefact.

  1. UNCERTAINTY IS UNDERSTATED.
     The original submission reported "±" values that were the standard
     deviation of per-fold AUROC across repeated cross-validation. Over
     repeated folds that reuse the same ~110 patients this is not a confidence
     interval and it understates uncertainty substantially. A naive bootstrap
     of the *pooled* out-of-fold predictions is equally wrong: after R repeats
     of 5-fold CV each patient contributes R rows to the pooled vector, so
     resampling rows treats one patient as R independent observations and
     shrinks the interval by roughly sqrt(R).
     → ESTIMAND (cv_estimands.py): the metric is computed on each CV repeat's
       complete out-of-fold vector (one prediction per patient per repeat)
       and averaged over repeats; the 95% CI is a PATIENT-LEVEL CLUSTER
       bootstrap in which all R predictions of a resampled patient move
       together. Averaging each patient's R probabilities into one number
       before computing AUROC — an earlier draft of this module — is NOT used
       for any performance number: it is an ensemble estimand that is
       optimistic for models with signal and severely pessimistic for
       near-constant ones (the held-out-outcome artefact described in
       cv_estimands.py; the clinical model dropped from 0.61 to 0.41).
       `pool_oof_by_patient` survives only as a per-patient RISK SCORE
       (tertile grouping fallback), never as a performance estimand.
       See SECTION 1.

  2. NO CALIBRATION REPORTING.
     → Reliability curves, calibration slope and intercept, and Brier score
       with a patient-level bootstrap CI, for the whole cohort and each arm.
       See SECTION 2.

  3. NO FEATURE-SELECTION STABILITY.
     → Selection frequency across folds with Wilson confidence intervals, and
       fold-wise stability of the fusion layer's modality weights.
       See SECTION 3.

  4. MODEL COMPARISONS WERE NOT TESTED, AND OVERFITTING RISK WAS NOT
     QUANTIFIED PER FOLD.
     → Paired tests on shared per-patient predictions (DeLong and paired
       bootstrap), and per-fold pCR event counts with realised
       events-per-variable. See SECTIONS 4 and 5.

  5. THE TREATMENT-SELECTION SCHEME WAS PRESENTED AS A DECISION TOOL.
     → Explicit pre-specified group rules, per-group and per-arm patient and
       event counts, and a formal treatment-by-group interaction test. These
       groups were derived and evaluated in the same cohort; the interaction
       term, not the separately elevated pCR rates, is what any
       differential-benefit reading rests on. See SECTION 6.

  6. THE RECURRENCE ANALYSIS WAS UNDERPOWERED AND UNCORRECTED.
     → Kaplan-Meier, log-rank and univariable Cox with Benjamini-Hochberg
       correction and explicit event counts, labelled exploratory throughout.
       See SECTION 7.

USAGE
-----
  python3 revision_analyses.py \\
      --results_dir ./results \\
      --out_dir     ./report \\
      --data_path   ./data/clin_multiomics_curated_metrics_PREDIX_HER2_new.txt

  # Survival section additionally needs follow-up columns:
  #   --survival_time_col DFS_months --survival_event_col DFS_event

OUTPUTS
-------
  {out_dir}/tables/revision/
    revision_performance_CI.xlsx        repeat-mean metrics, patient cluster-bootstrap CIs
    revision_calibration.xlsx           slope / intercept / Brier / reliability (per repeat)
    revision_stability.xlsx             selection frequency + modality weights
    revision_model_comparisons.xlsx     paired cluster bootstrap + per-repeat DeLong
    revision_epv_per_fold.xlsx          per-fold events and realised EPV
    revision_biomarker_groups.xlsx      group rules, counts, interaction test
    revision_survival_exploratory.xlsx  KM / log-rank / Cox with FDR
  {out_dir}/figures/revision/
    revfig01_calibration.pdf
    revfig02_selection_stability.pdf
    revfig03_epv_per_fold.pdf
    revfig04_biomarker_groups.pdf
    revfig05_survival_exploratory.pdf   (unused — survival dropped, no data)
    revfig07_model_comparisons.pdf      AUROC forest + paired ΔAUROC vs fused
    revfig08_fusion_weights.pdf         per-fold fusion weights + selection rate
  (revfig06_external_validation.pdf is written by external_validation.py)

DEPENDENCIES
------------
numpy, pandas, scipy, scikit-learn, matplotlib, openpyxl. No survival package
is required: Kaplan-Meier, the log-rank test and Breslow-tied Cox estimation
are implemented here so the repository has no dependency the reviewers cannot
install.
"""

import argparse
import pickle
import warnings
from pathlib import Path
from collections import defaultdict, Counter

warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from scipy import stats
from scipy.optimize import minimize
from sklearn.metrics import roc_auc_score, average_precision_score, brier_score_loss
from sklearn.linear_model import LogisticRegression

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Repeat-aware, patient-clustered estimands (shared with generate_report.py and
# external_validation.py so every quoted number rests on ONE definition).
import cv_estimands as CE


# =============================================================================
# CONSTANTS — kept identical to generate_report.py / multimodal_pcr_pipeline.py
# =============================================================================
SCENARIOS  = ["Global", "DHP", "T-DM1"]
EXP_MAP    = {"Global": "global", "DHP": "dhp", "T-DM1": "tdm1"}
UNIMODALS  = ["Clin", "RNA", "DNA", "Prot", "WSI"]
FUSED      = "Fused_ElasticNet"
ALL_MODELS = UNIMODALS + [FUSED]

MOD_COLOR  = {"Clin": "#4e79a7", "RNA": "#f28e2b", "DNA": "#e15759",
              "Prot": "#76b7b2", "WSI": "#59a14f", FUSED: "#6a1f6a"}
SC_COL     = {"Global": "#333333", "DHP": "#2166ac", "T-DM1": "#d6604d"}

# Pre-specified stability thresholds. A feature is described as a robust
# determinant only when its selection frequency exceeds these values. They are
# fixed in advance, not chosen after inspecting the frequencies, and they match
# the pipeline defaults (--stability_thresh_global / --stability_thresh_arm).
STABILITY_THRESH = {"Global": 0.60, "DHP": 0.50, "T-DM1": 0.50}

N_BOOT   = 2000    # patient-level bootstrap resamples
BOOT_SEED = 20240517
N_CAL_BINS = 10    # reliability-curve bins

plt.rcParams.update({
    "font.family": "sans-serif", "font.sans-serif": ["Arial", "DejaVu Sans"],
    "font.size": 9, "axes.titlesize": 10, "axes.titleweight": "bold",
    "axes.labelsize": 9, "xtick.labelsize": 8, "ytick.labelsize": 8,
    "legend.fontsize": 8, "legend.framealpha": 0.85,
    "axes.spines.top": False, "axes.spines.right": False,
    "axes.linewidth": 0.8, "grid.linewidth": 0.5, "grid.alpha": 0.4,
    "savefig.bbox": "tight", "savefig.dpi": 300,
})


# =============================================================================
# SECTION 0 — LOADING
# =============================================================================

def load_results(results_dir: Path) -> dict:
    """
    Load the discovery and consensus PKLs for every scenario that exists.

    Returns
    -------
    {scenario: {"discovery": {...} | None, "consensus": {...} | None}}
    Scenarios with neither artefact present are omitted.
    """
    out = {}
    for sc, exp in EXP_MAP.items():
        entry = {"discovery": None, "consensus": None}
        disc = results_dir / exp / f"{exp}_elasticnet_results.pkl"
        cons = results_dir / exp / f"{exp}_consensus_eval.pkl"
        if disc.exists():
            with open(disc, "rb") as f:
                entry["discovery"] = pickle.load(f)
        if cons.exists():
            with open(cons, "rb") as f:
                entry["consensus"] = pickle.load(f)
        if entry["discovery"] is not None or entry["consensus"] is not None:
            out[sc] = entry
    if not out:
        raise FileNotFoundError(
            f"No pipeline PKL files found under {results_dir}. "
            "Run multimodal_pcr_pipeline.py first.")
    return out


# =============================================================================
# SECTION 1 — PATIENT-LEVEL POOLING AND BOOTSTRAP CONFIDENCE INTERVALS
# =============================================================================
# This section is the direct answer to "the reported AUCs exceeding 80% with
# narrow confidence intervals are implausible for a cohort of this size".
#
# The narrowness had two separate causes and both are fixed here:
#   (a) the reported "±" was a standard deviation of per-fold AUROC, not a
#       confidence interval at all;
#   (b) any bootstrap taken over the pooled out-of-fold vector resamples ROWS,
#       and after R repeats of K-fold CV each patient owns R rows. The
#       resulting interval is approximately sqrt(R) times too narrow.
#
# The fix is to make the PATIENT the resampling unit while keeping the
# estimand a single-model quantity: the metric is computed on each repeat's
# complete out-of-fold vector and averaged over repeats, and the bootstrap
# resamples patients WITH all their R predictions (cluster bootstrap). This is
# implemented in cv_estimands.py and used by every performance table below.
#
# An earlier draft averaged each patient's R probabilities into one number
# first. That is NOT the same thing: it scores a 200-model ensemble, and for a
# near-constant model it is dominated by the held-out-outcome artefact (a
# held-out pCR patient's fold model was trained on one fewer event, so its
# prediction is shifted down by a hair in every one of its 200 repeats).
# `pool_oof_by_patient` is therefore kept ONLY as a per-patient risk score for
# the tertile-grouping fallback in SECTION 6, never for a performance metric.
# =============================================================================

def _repeat_matrix(blob, source, model):
    """
    (repeats x patients) out-of-fold matrix for one model, from either the
    consensus-eval PKL (source == "consensus") or the discovery PKL
    (source == "discovery"). Returns None when the model is absent.
    """
    if blob is None:
        return None
    if source == "consensus":
        folds = blob.get("folds", [])
        if not folds:
            return None
        if model != FUSED and model not in folds[0].get("unimodal_y_pred", {}):
            return None
        return CE.consensus_repeat_matrix(blob, model, fused_name=FUSED)
    folds = blob.get(model, [])
    if not folds:
        return None
    return CE.discovery_repeat_matrix(folds)


def pool_oof_by_patient(folds, pred_key="y_pred", label_key="y_test",
                        pid_key="test_pids"):
    """
    Collapse repeated out-of-fold predictions to one value per patient.

    USE ONLY AS A RISK SCORE (e.g. tertile grouping). Do NOT compute AUROC or
    any other performance metric on the returned vector — see the section
    banner above and cv_estimands.py for why.

    Every outer fold contributes predictions for the patients in its test set.
    With R repeats of K-fold cross-validation each patient appears in exactly R
    outer test sets, so the pooled vector contains R predictions per patient.
    Averaging them yields one prediction per patient, which is what makes a
    patient-level bootstrap meaningful.

    Parameters
    ----------
    folds     : list of fold dicts from the pipeline PKL.
    pred_key  : key holding this fold's test predictions.
    label_key : key holding this fold's test labels.
    pid_key   : key holding this fold's test patient IDs. When absent (PKLs
                written before the revision), the function falls back to
                'test_idx'. If neither is present it raises, because silently
                pooling without patient identity is exactly the error this
                function exists to prevent.

    Returns
    -------
    pid   : np.ndarray (n_patients,)  patient identifiers, sorted
    y     : np.ndarray (n_patients,)  outcome per patient
    p     : np.ndarray (n_patients,)  mean predicted probability per patient
    n_rep : np.ndarray (n_patients,)  number of predictions averaged
    """
    if not folds:
        return (np.array([]), np.array([]), np.array([]), np.array([]))

    key = None
    for candidate in (pid_key, "test_pids", "test_idx"):
        if candidate and candidate in folds[0]:
            key = candidate
            break
    if key is None:
        raise KeyError(
            "Fold dicts carry no patient identifier ('test_pids' or "
            "'test_idx'). Patient-level pooling is impossible without one. "
            "Re-run multimodal_pcr_pipeline.py with the revised pipeline, "
            "which records test_pids on every fold.")

    sums   = defaultdict(float)
    counts = defaultdict(int)
    labels = {}

    for fd in folds:
        pids = np.asarray(fd[key]).ravel()
        preds = np.asarray(fd[pred_key], dtype=float).ravel()
        ys    = np.asarray(fd[label_key], dtype=float).ravel()
        if not (len(pids) == len(preds) == len(ys)):
            raise ValueError(
                f"Fold {fd.get('fold_idx')} has mismatched lengths: "
                f"{len(pids)} ids, {len(preds)} predictions, {len(ys)} labels.")
        for pid, pr, yv in zip(pids, preds, ys):
            if np.isnan(pr):
                continue
            sums[int(pid)]   += float(pr)
            counts[int(pid)] += 1
            labels[int(pid)]  = float(yv)

    pids_sorted = np.array(sorted(counts.keys()), dtype=np.int64)
    p     = np.array([sums[i] / counts[i] for i in pids_sorted])
    n_rep = np.array([counts[i] for i in pids_sorted], dtype=int)
    y     = np.array([labels[i] for i in pids_sorted])
    return pids_sorted, y, p, n_rep


def pool_consensus_by_patient(cons, model):
    """
    Patient-level pooling for the consensus-evaluation PKL, whose fold dicts
    store unimodal predictions in a nested dict rather than under 'y_pred'.
    RISK SCORE ONLY (see pool_oof_by_patient); performance metrics use
    cv_estimands.consensus_repeat_matrix.

    The identifier key is resolved ONCE from the first fold, not per fold: a
    per-fold `f.get("test_pids", f.get("test_idx"))` would silently mix
    patient IDs with positional indices if folds from different pipeline
    versions were ever merged, and numerically colliding values would then
    average predictions of different patients. With a single key, a fold
    lacking it raises KeyError — the loud failure this module prefers.
    """
    folds = cons.get("folds", [])
    if not folds:
        return (np.array([]),) * 4
    key = "test_pids" if "test_pids" in folds[0] else "test_idx"
    if model == FUSED:
        shim = [{"test_pids": f[key],
                 "y_test": f["y_test"],
                 "y_pred": f["fused_y_pred"],
                 "fold_idx": f.get("fold_idx")} for f in folds]
    else:
        shim = [{"test_pids": f[key],
                 "y_test": f["y_test"],
                 "y_pred": f["unimodal_y_pred"][model],
                 "fold_idx": f.get("fold_idx")} for f in folds]
    return pool_oof_by_patient(shim)


def _safe_auroc(y, p):
    if len(np.unique(y)) < 2:
        return np.nan
    return float(roc_auc_score(y, p))


def _safe_auprc(y, p):
    if len(np.unique(y)) < 2:
        return np.nan
    return float(average_precision_score(y, p))


def _safe_brier(y, p):
    if len(y) == 0:
        return np.nan
    return float(brier_score_loss(y, np.clip(p, 0.0, 1.0)))


METRIC_FNS = {"AUROC": _safe_auroc, "AUPRC": _safe_auprc, "Brier": _safe_brier}


def bootstrap_metric_ci(y, p, metric="AUROC", n_boot=None, seed=None,
                        ci=0.95, stratified=True):
    """
    Patient-level bootstrap confidence interval.

    n_boot/seed default to None and resolve to the module globals AT CALL
    TIME. Def-time defaults (n_boot=N_BOOT) freeze the import-time values,
    which silently ignores the --n_boot/--seed CLI overrides main() applies
    by rebinding the globals — while the Excel notes print the new values.

    One row of (y, p) MUST be one patient with ONE prediction each — an
    external cohort predicted once by a frozen model, for instance. For
    repeated-CV out-of-fold predictions use cv_estimands.bootstrap_repeat_
    metric_ci instead; do NOT feed this function patient-averaged CV
    probabilities (see the SECTION 1 banner).

    Stratified resampling (default) draws separately from the event and
    non-event strata, preserving the observed number of pCR events in every
    resample. At 46 events in 110 patients an unstratified bootstrap produces
    a non-trivial number of resamples with extreme event counts, which inflates
    the variance of the AUROC estimate for reasons unrelated to the model.
    Stratified resampling is the standard choice for this situation and is
    pre-specified here.

    Returns
    -------
    dict with point estimate, ci_low, ci_high, se, n, n_events, n_boot_valid.
    """
    if n_boot is None: n_boot = N_BOOT
    if seed is None: seed = BOOT_SEED
    y = np.asarray(y, dtype=float)
    p = np.asarray(p, dtype=float)
    fn = METRIC_FNS[metric]

    n = len(y)
    result = {"metric": metric, "estimate": np.nan, "ci_low": np.nan,
              "ci_high": np.nan, "se": np.nan, "n": n,
              "n_events": int(np.nansum(y)), "n_boot_valid": 0}
    if n == 0:
        return result

    result["estimate"] = fn(y, p)
    rng = np.random.default_rng(seed)

    idx_pos = np.where(y == 1)[0]
    idx_neg = np.where(y == 0)[0]

    boots = np.empty(n_boot, dtype=float)
    for b in range(n_boot):
        if stratified and len(idx_pos) > 0 and len(idx_neg) > 0:
            take = np.concatenate([
                rng.choice(idx_pos, size=len(idx_pos), replace=True),
                rng.choice(idx_neg, size=len(idx_neg), replace=True)])
        else:
            take = rng.integers(0, n, size=n)
        boots[b] = fn(y[take], p[take])

    boots = boots[np.isfinite(boots)]
    if len(boots) == 0:
        return result

    alpha = (1.0 - ci) / 2.0
    result["ci_low"]       = float(np.percentile(boots, alpha * 100))
    result["ci_high"]      = float(np.percentile(boots, (1 - alpha) * 100))
    result["se"]           = float(np.std(boots, ddof=1))
    result["n_boot_valid"] = int(len(boots))
    return result


def format_ci(res, fmt="{:.3f}"):
    """Render a bootstrap result as 'estimate [low-high]'."""
    if not np.isfinite(res.get("estimate", np.nan)):
        return "—"
    if not np.isfinite(res.get("ci_low", np.nan)):
        return fmt.format(res["estimate"])
    return (f"{fmt.format(res['estimate'])} "
            f"[{fmt.format(res['ci_low'])}–{fmt.format(res['ci_high'])}]")


# =============================================================================
# SECTION 2 — CALIBRATION
# =============================================================================

def calibration_metrics(y, p, n_bins=N_CAL_BINS, n_boot=None, seed=None):
    """
    Calibration assessment for ONE prediction per patient (e.g. an external
    cohort scored once by a frozen model). For repeated-CV out-of-fold
    predictions the reported calibration comes from
    cv_estimands.calibration_repeat_summary (per repeat, averaged) — do not
    pass patient-averaged CV probabilities here (see the SECTION 1 banner).

    Reported quantities
    -------------------
    slope, intercept
        From the logistic recalibration model logit(P(y=1)) = a + b*logit(p).
        Slope 1 and intercept 0 indicate perfect calibration. Slope below 1
        means the predictions are too extreme (over-fitted); slope above 1
        means they are too compressed. Intercept captures calibration-in-the-
        large: whether the model's average predicted risk matches the observed
        event rate.

    Brier score with patient-level bootstrap CI
        Proper scoring rule combining calibration and discrimination.

    ECE
        Expected calibration error: the average absolute gap between predicted
        and observed risk across equal-count bins, weighted by bin size.

    reliability
        Per-bin mean predicted risk, observed risk, and count — the data for
        the reliability curve. Bins are equal-count (quantile) rather than
        equal-width because with 110 patients equal-width bins leave several
        bins empty or holding a single patient.

    Note on interpretation: the probabilities entering this function are
    Platt-recalibrated within the cross-validation, so a slope near 1 confirms
    the recalibration transferred to held-out patients rather than
    demonstrating that the underlying classifier was calibrated to begin with.

    n_boot/seed: None resolves to the module globals at call time so the
    --n_boot/--seed CLI overrides actually apply (see bootstrap_metric_ci).
    """
    if n_boot is None: n_boot = N_BOOT
    if seed is None: seed = BOOT_SEED
    y = np.asarray(y, dtype=float)
    p = np.clip(np.asarray(p, dtype=float), 1e-6, 1 - 1e-6)

    out = {"n": len(y), "n_events": int(np.nansum(y)),
           "slope": np.nan, "intercept": np.nan,
           "slope_ci": (np.nan, np.nan), "intercept_ci": (np.nan, np.nan),
           "brier": np.nan, "brier_ci": (np.nan, np.nan),
           "ece": np.nan, "observed_rate": np.nan, "mean_predicted": np.nan,
           "reliability": pd.DataFrame()}

    if len(y) < 10 or len(np.unique(y)) < 2:
        return out

    logit_p = np.log(p / (1 - p))

    def _fit_slope_intercept(yy, lp):
        """Logistic recalibration; returns (slope, intercept)."""
        try:
            lr = LogisticRegression(penalty=None, solver="lbfgs", max_iter=1000)
            lr.fit(lp.reshape(-1, 1), yy)
            return float(lr.coef_[0][0]), float(lr.intercept_[0])
        except Exception:
            return np.nan, np.nan

    slope, intercept = _fit_slope_intercept(y, logit_p)
    out["slope"], out["intercept"] = slope, intercept
    out["brier"] = _safe_brier(y, p)
    out["observed_rate"]  = float(np.mean(y))
    out["mean_predicted"] = float(np.mean(p))

    # ── Patient-level bootstrap for slope, intercept and Brier ───────────────
    rng = np.random.default_rng(seed)
    idx_pos = np.where(y == 1)[0]
    idx_neg = np.where(y == 0)[0]
    sl_b, ic_b, br_b = [], [], []
    for _ in range(n_boot):
        take = np.concatenate([
            rng.choice(idx_pos, size=len(idx_pos), replace=True),
            rng.choice(idx_neg, size=len(idx_neg), replace=True)])
        s, i = _fit_slope_intercept(y[take], logit_p[take])
        if np.isfinite(s):
            sl_b.append(s)
            ic_b.append(i)
        br_b.append(_safe_brier(y[take], p[take]))

    def _pct(arr):
        arr = np.asarray([v for v in arr if np.isfinite(v)], dtype=float)
        if len(arr) == 0:
            return (np.nan, np.nan)
        return (float(np.percentile(arr, 2.5)), float(np.percentile(arr, 97.5)))

    out["slope_ci"]     = _pct(sl_b)
    out["intercept_ci"] = _pct(ic_b)
    out["brier_ci"]     = _pct(br_b)

    # ── Reliability curve on equal-count bins ────────────────────────────────
    n_bins_eff = int(min(n_bins, max(2, len(y) // 8)))
    ranks = np.argsort(p)
    bin_of = np.empty(len(p), dtype=int)
    bin_of[ranks] = (np.arange(len(p)) * n_bins_eff) // len(p)

    rows = []
    abs_gap_weighted = 0.0
    for b in range(n_bins_eff):
        m = bin_of == b
        if not m.any():
            continue
        pred_mean = float(np.mean(p[m]))
        obs_rate  = float(np.mean(y[m]))
        cnt       = int(m.sum())
        # Wilson interval on the observed rate so the reliability plot shows
        # how much of any apparent miscalibration is bin-level noise.
        lo, hi = wilson_ci(int(y[m].sum()), cnt)
        rows.append({"bin": b + 1, "n": cnt, "n_events": int(y[m].sum()),
                     "mean_predicted": pred_mean, "observed": obs_rate,
                     "obs_ci_low": lo, "obs_ci_high": hi})
        abs_gap_weighted += cnt * abs(pred_mean - obs_rate)

    out["reliability"] = pd.DataFrame(rows)
    out["ece"] = float(abs_gap_weighted / len(y)) if len(y) else np.nan
    return out


def wilson_ci(successes: int, n: int, z: float = 1.959963985):
    """
    Wilson score interval for a binomial proportion.

    Used for selection frequencies and for the observed rate inside each
    reliability bin. Preferred over the normal approximation because both
    applications involve small counts and proportions near 0 or 1, where the
    normal interval can extend outside [0, 1].
    """
    if n == 0:
        return (np.nan, np.nan)
    phat = successes / n
    denom = 1 + z ** 2 / n
    centre = (phat + z ** 2 / (2 * n)) / denom
    half = (z * np.sqrt(phat * (1 - phat) / n + z ** 2 / (4 * n ** 2))) / denom
    return (float(max(0.0, centre - half)), float(min(1.0, centre + half)))


# =============================================================================
# SECTION 3 — FEATURE-SELECTION AND MODALITY-WEIGHT STABILITY
# =============================================================================

def selection_frequency(folds, threshold):
    """
    Per-feature selection frequency across outer folds, with Wilson CIs.

    A feature's frequency is the fraction of outer folds in which it appeared
    in that fold's winner signature. The pre-specified stability threshold
    decides whether the feature may be described as a robust determinant; it
    is applied to the point estimate, and the Wilson lower bound is reported
    alongside so a frequency that only just clears the threshold is visible as
    such.

    Two denominators matter and both are reported:
      n_folds_total     — every outer fold.
      n_folds_eligible  — folds in which the feature survived that fold's
                          preprocessing and could therefore have been picked
                          (read from the fold's "candidate_features" key).
                          A feature culled by a fold's near-zero-variance or
                          correlation filter never had the opportunity to be
                          selected, and scoring it as "not selected" in that
                          fold understates its stability. PKLs from pipeline
                          versions that did not record the candidate pool
                          fall back to the all-folds denominator.
    """
    winner_folds = [f for f in folds
                    if f.get("winner_clf", "") not in ("", "none")
                    and f.get("winner_signature")]
    n_total = len(winner_folds)
    if n_total == 0:
        return pd.DataFrame()

    sel_count = Counter()
    eligible  = Counter()
    imp_vals  = defaultdict(list)

    # The candidate pool must come from "candidate_features" (the post-
    # preprocessing pool the pipeline records per fold). The fold dict's
    # "features" key holds the WINNER SIGNATURE — using it as the pool makes
    # eligible == selected for every feature, so selection_freq_eligible
    # would be 1.0 across the board and the whole stability deliverable
    # vacuous. Old PKLs lacking the key fall back to the all-folds
    # denominator (conservative), never to selected/selected.
    any_pool = any(f.get("candidate_features") for f in winner_folds)
    # Folds lacking the key (old-PKL folds mixed into a newer file) count
    # into EVERY feature's eligible denominator. Without this, a no-pool
    # fold's eligible set degenerates to its own signature and the
    # selected/selected = 1.0 vacuity returns on a per-fold basis.
    n_no_pool = sum(1 for f in winner_folds
                    if not f.get("candidate_features"))

    for f in winner_folds:
        sig = set(f.get("winner_signature", []))
        w    = f.get("winner_clf", "")
        # True importance magnitudes; "inner_importance" holds percentile
        # ranks (the pipeline's own warning says not to present those as
        # importances). Fall back only for old PKLs.
        imp = (f.get("inner_importance_magnitude")
               or f.get("inner_importance", {})).get(w, {})
        if any_pool and f.get("candidate_features"):
            pool = set(f["candidate_features"]) | sig
            for feat in pool:
                eligible[feat] += 1
        for feat in sig:
            sel_count[feat] += 1
            imp_vals[feat].append(abs(float(imp.get(feat, 0.0))))

    rows = []
    for feat, cnt in sel_count.items():
        elig = max((eligible.get(feat, 0) + n_no_pool) if any_pool
                   else n_total, 1)
        freq_total = cnt / n_total
        freq_elig  = cnt / elig
        lo, hi = wilson_ci(cnt, elig)
        rows.append({
            "feature": feat,
            "n_selected": cnt,
            "n_folds_total": n_total,
            "n_folds_eligible": elig,
            "selection_freq": freq_total,
            "selection_freq_eligible": freq_elig,
            "wilson_low": lo,
            "wilson_high": hi,
            "mean_importance_when_selected": (
                float(np.mean(imp_vals[feat])) if imp_vals[feat] else np.nan),
            "stability_threshold": threshold,
            "stable": bool(freq_elig >= threshold),
            "stable_ci_supported": bool(lo >= threshold),
        })

    # Deterministic row order. `rows` is built from Counter insertion order,
    # which follows iteration over a set of feature-name strings and therefore
    # varies between processes (Python randomises string hashing per run).
    # With a non-stable sort the tied rows — and at these frequencies many
    # features tie at 1.0 — then come out in a different order on every run,
    # so two identical analyses produce workbooks that differ by a row
    # permutation. Sorting on the feature name as the final key, with a stable
    # kind, makes the sheet byte-reproducible without changing any value.
    df = pd.DataFrame(rows).sort_values(
        ["selection_freq_eligible", "selection_freq", "feature"],
        ascending=[False, False, True], kind="mergesort").reset_index(drop=True)
    return df


def modality_weight_stability(folds):
    """
    Fold-wise stability of the late-fusion modality weights.

    The fusion layer is a penalised (elastic-net) logistic model over the five
    calibrated modality probability streams. Its L1 component sets
    non-contributing modalities to exactly zero, so the natural stability
    statistics are how often each modality received a non-zero weight and
    whether its sign was consistent — not merely the mean coefficient, which
    can be near zero either because the modality is unhelpful or because its
    sign flips across folds.
    """
    rows = []
    if not folds:
        return pd.DataFrame()

    weights = defaultdict(list)
    for f in folds:
        mw = f.get("modality_weights", {})
        for mod, w in mw.items():
            weights[mod].append(float(w))

    n_folds = len(folds)
    for mod, vals in weights.items():
        arr = np.asarray(vals, dtype=float)
        nz = np.abs(arr) > 1e-6
        n_nonzero = int(nz.sum())
        sign_pos = int((arr > 1e-6).sum())
        sign_neg = int((arr < -1e-6).sum())
        dominant = max(sign_pos, sign_neg)
        lo, hi = wilson_ci(n_nonzero, n_folds)
        rows.append({
            "modality": mod,
            "n_folds": n_folds,
            "n_nonzero": n_nonzero,
            "selection_rate": n_nonzero / n_folds if n_folds else np.nan,
            "selection_rate_ci_low": lo,
            "selection_rate_ci_high": hi,
            "mean_weight": float(np.mean(arr)),
            "median_weight": float(np.median(arr)),
            "sd_weight": float(np.std(arr, ddof=1)) if len(arr) > 1 else 0.0,
            "weight_p2.5": float(np.percentile(arr, 2.5)),
            "weight_p97.5": float(np.percentile(arr, 97.5)),
            # Sign consistency among folds where the modality was retained.
            "sign_consistency": (dominant / n_nonzero) if n_nonzero else np.nan,
        })
    if not rows:
        # Foreign/old PKLs whose fused folds lack modality_weights: sorting
        # an empty frame on a named column raises KeyError.
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values(
        "selection_rate", ascending=False).reset_index(drop=True)


# =============================================================================
# SECTION 4 — PAIRED MODEL COMPARISON
# =============================================================================
# Every model in this study is evaluated on the same patients, so comparisons
# must be paired. Two complementary tests are reported: DeLong, which is the
# standard analytic test for two correlated ROC curves, and a paired patient-
# level bootstrap, which makes no distributional assumption and extends to
# metrics other than AUROC.
# =============================================================================

def _fast_delong_structural(y, preds):
    """
    Structural components of the fast DeLong algorithm (Sun & Xu, 2014).

    preds : (n_models, n_patients)
    Returns (aucs, covariance_matrix).
    """
    y = np.asarray(y)
    order = np.argsort(-y, kind="mergesort")   # positives first
    preds = np.asarray(preds, dtype=float)[:, order]
    m = int(np.sum(y == 1))
    n = len(y) - m
    k = preds.shape[0]

    if m == 0 or n == 0:
        return np.full(k, np.nan), np.full((k, k), np.nan)

    def _midrank(x):
        srt = np.argsort(x)
        xs = x[srt]
        n_x = len(x)
        r = np.empty(n_x, dtype=float)
        i = 0
        while i < n_x:
            j = i
            while j < n_x - 1 and xs[j + 1] == xs[i]:
                j += 1
            r[i:j + 1] = 0.5 * (i + j) + 1
            i = j + 1
        out = np.empty(n_x, dtype=float)
        out[srt] = r
        return out

    tx = np.empty((k, m), dtype=float)
    ty = np.empty((k, n), dtype=float)
    tz = np.empty((k, m + n), dtype=float)
    for r in range(k):
        tx[r] = _midrank(preds[r, :m])
        ty[r] = _midrank(preds[r, m:])
        tz[r] = _midrank(preds[r])

    aucs = (tz[:, :m].sum(axis=1) / m / n
            - float(m + 1.0) / 2.0 / n)
    v01 = (tz[:, :m] - tx) / n
    v10 = 1.0 - (tz[:, m:] - ty) / m
    s01 = np.cov(v01)
    s10 = np.cov(v10)
    if k == 1:
        s01 = np.array([[float(s01)]])
        s10 = np.array([[float(s10)]])
    cov = s01 / m + s10 / n
    return aucs, cov


def delong_test(y, p1, p2):
    """
    DeLong test for two correlated ROC curves evaluated on the same patients.

    Returns dict with auc1, auc2, delta, se, z, p_value.
    Inputs must be patient-level (one row per patient).
    """
    y = np.asarray(y, dtype=float)
    out = {"auc1": np.nan, "auc2": np.nan, "delta": np.nan,
           "se": np.nan, "z": np.nan, "p_value": np.nan}
    # Each class needs >= 2 members: with a single positive (or negative),
    # np.cov with ddof=1 yields NaN, `var <= 0` is False for NaN, and the
    # function would return se=nan / p=nan unflagged.
    if np.sum(y == 1) < 2 or np.sum(y == 0) < 2 or len(y) < 5:
        return out

    aucs, cov = _fast_delong_structural(y, np.vstack([p1, p2]))
    if not np.all(np.isfinite(aucs)):
        return out

    out["auc1"], out["auc2"] = float(aucs[0]), float(aucs[1])
    delta = float(aucs[0] - aucs[1])
    var = float(cov[0, 0] + cov[1, 1] - 2 * cov[0, 1])
    out["delta"] = delta
    if not np.isfinite(var) or var <= 0:
        # Identical predictions, or numerically degenerate covariance.
        out["se"] = 0.0
        out["z"] = 0.0
        out["p_value"] = 1.0
        return out
    se = float(np.sqrt(var))
    z = delta / se
    out["se"] = se
    out["z"] = float(z)
    out["p_value"] = float(2 * stats.norm.sf(abs(z)))
    return out


def paired_bootstrap_delta(y, p1, p2, metric="AUROC",
                           n_boot=None, seed=None):
    """
    Paired patient-level bootstrap of the difference in a metric.

    The SAME resampled patient indices are applied to both models, which is
    what makes the comparison paired and removes the between-model correlation
    from the interval. The two-sided p-value is the standard bootstrap
    inversion: twice the smaller tail proportion of resampled differences
    lying on the wrong side of zero.

    n_boot/seed: None resolves to the module globals at call time so the
    --n_boot/--seed CLI overrides actually apply (see bootstrap_metric_ci).
    """
    if n_boot is None: n_boot = N_BOOT
    if seed is None: seed = BOOT_SEED
    y = np.asarray(y, dtype=float)
    p1 = np.asarray(p1, dtype=float)
    p2 = np.asarray(p2, dtype=float)
    fn = METRIC_FNS[metric]

    out = {"metric": metric, "delta": np.nan, "ci_low": np.nan,
           "ci_high": np.nan, "p_value": np.nan, "n": len(y),
           "n_events": int(np.nansum(y))}
    if len(y) == 0 or len(np.unique(y)) < 2:
        return out

    out["delta"] = fn(y, p1) - fn(y, p2)

    rng = np.random.default_rng(seed)
    idx_pos = np.where(y == 1)[0]
    idx_neg = np.where(y == 0)[0]
    deltas = np.empty(n_boot, dtype=float)
    for b in range(n_boot):
        take = np.concatenate([
            rng.choice(idx_pos, size=len(idx_pos), replace=True),
            rng.choice(idx_neg, size=len(idx_neg), replace=True)])
        deltas[b] = fn(y[take], p1[take]) - fn(y[take], p2[take])

    deltas = deltas[np.isfinite(deltas)]
    if len(deltas) == 0:
        return out

    out["ci_low"]  = float(np.percentile(deltas, 2.5))
    out["ci_high"] = float(np.percentile(deltas, 97.5))
    prop_le = float(np.mean(deltas <= 0))
    prop_ge = float(np.mean(deltas >= 0))
    # Guard against p = 0, which a finite bootstrap cannot support.
    p = 2 * min(prop_le, prop_ge)
    out["p_value"] = float(min(1.0, max(p, 1.0 / len(deltas))))
    return out


def compare_models_repeat(y, P_by_model, reference=FUSED, seed=None):
    """
    Compare `reference` against every other model on the SAME patients and
    the SAME cross-validation repeats, using the repeat-aware estimand.

    P_by_model : {model: (R x n) out-of-fold matrix} — identical column
                 order (patients) and row order (repeats) for every model.
    Primary test: paired patient-level cluster bootstrap of
        delta = mean_r AUROC(y, P_ref[r]) - mean_r AUROC(y, P_alt[r]),
    the same patient resample applied to both models and every repeat.
    Secondary: DeLong's test run once per repeat (it needs one prediction per
    patient) and summarised across repeats (median p, fraction < 0.05). The
    verdict rests on the bootstrap interval only, and is deliberately
    conservative: an interval that includes zero is 'not distinguishable',
    however large the point difference.
    """
    if seed is None:
        seed = BOOT_SEED
    rows = []
    if reference not in P_by_model:
        return pd.DataFrame()
    P_ref = P_by_model[reference]
    for mod, P_alt in P_by_model.items():
        if mod == reference:
            continue
        bs = CE.paired_bootstrap_repeat_delta(P_ref, P_alt, y, metric="AUROC",
                                              n_boot=N_BOOT, seed=seed)
        dl = CE.per_repeat_test_summary(P_ref, P_alt, y, delong_test)
        crosses_zero = (not np.isfinite(bs["ci_low"])
                        or (bs["ci_low"] <= 0 <= bs["ci_high"]))
        if crosses_zero:
            verdict = "not distinguishable"
        elif bs["delta"] > 0:
            verdict = f"{reference} higher"
        else:
            verdict = f"{mod} higher"
        rows.append({
            "reference": reference,
            "comparator": mod,
            "AUROC_reference": bs["estimate_1"],
            "AUROC_comparator": bs["estimate_2"],
            "delta_AUROC": bs["delta"],
            "delta_CI_low": bs["ci_low"],
            "delta_CI_high": bs["ci_high"],
            "p_bootstrap": bs["p_value"],
            "DeLong_p_median_over_repeats": dl["p_median"],
            "DeLong_p_IQR": (f"{dl['p_q25']:.3g}–{dl['p_q75']:.3g}"
                             if np.isfinite(dl["p_q25"]) else "—"),
            "DeLong_frac_repeats_p<0.05": dl["frac_p_below_0.05"],
            "n_patients": bs["n"],
            "n_events": bs["n_events"],
            "n_cv_repeats": bs["n_repeats"],
            "verdict": verdict,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df["q_bootstrap_BH"] = benjamini_hochberg(df["p_bootstrap"].values)
    return df


def compare_models(y, preds_by_model, reference=FUSED):
    """
    Compare `reference` against every other model on shared patients, ONE
    prediction per patient (e.g. external cohorts). For repeated-CV
    out-of-fold predictions use `compare_models_repeat`.

    Reports DeLong and paired-bootstrap results side by side, plus a plain
    verdict column. The verdict is deliberately conservative: a comparison
    whose 95% interval includes zero is reported as 'not distinguishable',
    never as the reference being better, however large the point estimate.
    """
    rows = []
    if reference not in preds_by_model:
        return pd.DataFrame()
    p_ref = preds_by_model[reference]

    for mod, p_alt in preds_by_model.items():
        if mod == reference:
            continue
        dl = delong_test(y, p_ref, p_alt)
        bs = paired_bootstrap_delta(y, p_ref, p_alt, metric="AUROC")
        crosses_zero = (not np.isfinite(bs["ci_low"])
                        or (bs["ci_low"] <= 0 <= bs["ci_high"]))
        if crosses_zero:
            verdict = "not distinguishable"
        elif bs["delta"] > 0:
            verdict = f"{reference} higher"
        else:
            verdict = f"{mod} higher"
        rows.append({
            "reference": reference,
            "comparator": mod,
            "AUROC_reference": dl["auc1"],
            "AUROC_comparator": dl["auc2"],
            "delta_AUROC": bs["delta"],
            "delta_CI_low": bs["ci_low"],
            "delta_CI_high": bs["ci_high"],
            "p_bootstrap": bs["p_value"],
            "p_DeLong": dl["p_value"],
            "n_patients": bs["n"],
            "n_events": bs["n_events"],
            "verdict": verdict,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df["q_DeLong_BH"] = benjamini_hochberg(df["p_DeLong"].values)
        df["q_bootstrap_BH"] = benjamini_hochberg(df["p_bootstrap"].values)
    return df


def benjamini_hochberg(pvals):
    """Benjamini-Hochberg FDR adjustment; returns q-values in input order."""
    p = np.asarray(pvals, dtype=float)
    ok = np.isfinite(p)
    q = np.full(len(p), np.nan)
    if ok.sum() == 0:
        return q
    pv = p[ok]
    m = len(pv)
    order = np.argsort(pv)
    ranked = pv[order]
    q_sorted = ranked * m / np.arange(1, m + 1)
    q_sorted = np.minimum.accumulate(q_sorted[::-1])[::-1]
    q_ok = np.empty(m)
    q_ok[order] = np.minimum(q_sorted, 1.0)
    q[ok] = q_ok
    return q


# =============================================================================
# SECTION 5 — PER-FOLD EVENT COUNTS AND REALISED EVENTS-PER-VARIABLE
# =============================================================================

def epv_table(folds, scenario, model):
    """
    Per-fold pCR event counts and realised events-per-variable.

    The pipeline caps signature size at EPV = 5 during discovery. That cap is
    a design constraint; what matters for judging overfitting is the EPV
    actually realised, which differs from the cap because of the minimum-size
    floor and because features can be dropped by a fold's preprocessing after
    the cap was applied. Both the nominal cap and the realised value are
    therefore reported per fold.
    """
    rows = []
    for f in folds:
        n_sig = f.get("signature_size", np.nan)
        n_ev_train = f.get("n_events_train_expanded",
                           f.get("n_events_inner", np.nan))
        epv = f.get("epv_realized", np.nan)
        if not np.isfinite(epv) and np.isfinite(n_ev_train) and n_sig:
            epv = n_ev_train / max(n_sig, 1)
        rows.append({
            "scenario": scenario,
            "model": model,
            "fold": f.get("fold_idx", np.nan),
            "n_test": f.get("n_test", len(f.get("y_test", []))),
            "n_events_test": f.get(
                "n_events_test",
                int(np.nansum(np.asarray(f.get("y_test", []), float)))),
            "n_train_cc": f.get("n_train_cc", np.nan),
            "n_events_train_cc": f.get("n_events_train_cc", np.nan),
            "n_train_expanded": f.get("n_train_expanded", np.nan),
            "n_events_train_expanded": n_ev_train,
            "n_candidates_after_screen": f.get("n_candidates_outer", np.nan),
            "signature_size": n_sig,
            "epv_realized": epv,
            "epv_cap": 5,
            "winner_clf": f.get("winner_clf", ""),
        })
    return pd.DataFrame(rows)


def summarise_epv(df_epv):
    """Collapse the per-fold EPV table to one row per scenario x model."""
    if df_epv.empty:
        return pd.DataFrame()

    def _agg(g):
        return pd.Series({
            "n_folds": len(g),
            "median_n_events_test": g["n_events_test"].median(),
            "min_n_events_test": g["n_events_test"].min(),
            "max_n_events_test": g["n_events_test"].max(),
            "median_n_events_train": g["n_events_train_expanded"].median(),
            "median_signature_size": g["signature_size"].median(),
            "median_epv_realized": g["epv_realized"].median(),
            "min_epv_realized": g["epv_realized"].min(),
            "pct_folds_epv_below_5": float(
                100.0 * np.mean(g["epv_realized"] < 5)),
        })

    return (df_epv.groupby(["scenario", "model"], as_index=False)
            .apply(_agg, include_groups=False)
            .reset_index(drop=True))


# =============================================================================
# SECTION 6 — EXPLORATORY BIOMARKER GROUPS AND TREATMENT INTERACTION
# =============================================================================
# The groups below were derived and evaluated in the SAME cohort. They are
# reported as exploratory and hypothesis-generating. Nothing in this section
# supports a treatment-selection claim, and the code deliberately reports the
# interaction term rather than comparing per-group pCR rates across arms,
# because separately elevated pCR rates in the two arms do not establish
# differential benefit.
# =============================================================================

def assign_groups_from_rules(df, rules):
    """
    Assign patients to biomarker groups using explicit, pre-specified rules.

    `rules` is an ordered list of dicts:
        [{"name": "Group A",
          "expr": "RNA_TILs > 0.5 and DNA_ERBB2_CNA > 2",
          "description": "immune-high, HER2-amplified"}, ...]

    Rules are evaluated in order and the FIRST match wins, so the list defines
    a partition. Patients matching no rule are labelled 'Unassigned' and are
    reported rather than dropped — an unassigned fraction is itself a finding
    about how well the scheme covers the cohort.

    Expressions are evaluated with pandas.eval against the dataframe columns.
    Writing the rules as strings, in one place, is what makes them auditable:
    they go into the supplementary table exactly as the code executed them.
    """
    labels = pd.Series(["Unassigned"] * len(df), index=df.index, dtype=object)
    assigned = pd.Series(False, index=df.index)
    applied = []

    for rule in rules:
        name = rule["name"]
        expr = rule["expr"]
        try:
            mask = df.eval(expr)
            mask = mask.astype(bool) & (~assigned)
        except Exception as e:
            print(f"  [GROUPS] Rule '{name}' could not be evaluated "
                  f"({type(e).__name__}: {e}) — skipped.")
            applied.append({"group": name, "expr": expr, "n_matched": 0,
                            "status": f"ERROR: {e}"})
            continue
        labels[mask] = name
        assigned |= mask
        applied.append({"group": name, "expr": expr,
                        "description": rule.get("description", ""),
                        "n_matched": int(mask.sum()), "status": "applied"})

    n_unassigned = int((~assigned).sum())
    applied.append({"group": "Unassigned", "expr": "(no rule matched)",
                    "description": "patients matching none of the rules above",
                    "n_matched": n_unassigned, "status": "applied"})
    return labels, pd.DataFrame(applied)


def group_outcome_table(df, group_col, arm_col, outcome_col):
    """
    Patient and pCR event counts per group, overall and within each arm.

    Wilson intervals are attached to every pCR rate. With groups this small
    the point estimates are unstable, and showing the interval is what stops
    a 3/4 = 75% cell being read as a high-response subgroup.
    """
    rows = []
    for grp, g in df.groupby(group_col, dropna=False):
        for arm_label, sub in [("Overall", g)] + [
                (str(a), g[g[arm_col] == a]) for a in sorted(df[arm_col].dropna().unique())]:
            n = len(sub)
            if n == 0:
                rows.append({"group": grp, "arm": arm_label, "n": 0,
                             "n_pCR": 0, "pCR_rate": np.nan,
                             "pCR_ci_low": np.nan, "pCR_ci_high": np.nan})
                continue
            ev = int(np.nansum(sub[outcome_col].values))
            lo, hi = wilson_ci(ev, n)
            rows.append({"group": grp, "arm": arm_label, "n": n,
                         "n_pCR": ev, "pCR_rate": ev / n,
                         "pCR_ci_low": lo, "pCR_ci_high": hi})
    return pd.DataFrame(rows)


def _fit_logit_manual(X, y, max_iter=100, tol=1e-8, ridge=1e-8):
    """
    Unpenalised logistic regression by Newton-Raphson (IRLS).

    A tiny ridge term keeps the Hessian invertible under quasi-separation,
    which does occur in the smaller group x arm cells. Returns
    (log_likelihood, coefficients, standard_errors). The ridge is small enough
    (1e-8) not to move the log-likelihood materially, so the LRT remains valid,
    but it prevents an outright LinAlgError on a singular design.
    """
    Xd = np.hstack([np.ones((X.shape[0], 1)), X])
    beta = np.zeros(Xd.shape[1])
    for _ in range(max_iter):
        eta = Xd @ beta
        pr = 1.0 / (1.0 + np.exp(-np.clip(eta, -35, 35)))
        W = pr * (1 - pr)
        grad = Xd.T @ (y - pr) - ridge * beta
        H = -(Xd.T * W) @ Xd - ridge * np.eye(Xd.shape[1])
        try:
            step = np.linalg.solve(H, grad)
        except np.linalg.LinAlgError:
            step = np.linalg.lstsq(H, grad, rcond=None)[0]
        beta_new = beta - step
        if np.max(np.abs(beta_new - beta)) < tol:
            beta = beta_new
            break
        beta = beta_new

    eta = Xd @ beta
    pr = np.clip(1.0 / (1.0 + np.exp(-np.clip(eta, -35, 35))), 1e-12, 1 - 1e-12)
    ll = float(np.sum(y * np.log(pr) + (1 - y) * np.log(1 - pr)))
    W = pr * (1 - pr)
    try:
        cov = np.linalg.inv((Xd.T * W) @ Xd + ridge * np.eye(Xd.shape[1]))
        se = np.sqrt(np.clip(np.diag(cov), 0, np.inf))
    except np.linalg.LinAlgError:
        se = np.full(len(beta), np.nan)
    return ll, beta, se


def risk_tertile_groups(pid, y, p, arm, labels=("Low", "Intermediate", "High")):
    """
    Fallback grouping when neither the S-group columns nor explicit rules are
    available: tertiles of the model's own out-of-fold predicted probability.

    This is NOT a biomarker scheme and is not a substitute for one. It exists
    so the interaction machinery can be exercised and reported even before the
    manuscript's group definitions are wired in, and so the same test can be
    seen applied to a completely pre-specified partition rule.
    """
    q = np.quantile(p, [1 / 3, 2 / 3])
    grp = np.where(p <= q[0], labels[0],
                   np.where(p <= q[1], labels[1], labels[2]))
    return pd.DataFrame({"patient_id": pid, "pCR": y, "p_pred": p,
                         "arm": arm, "group": grp}), q


# -----------------------------------------------------------------------------
# The S1 / S2 / S3 step-wise classification (manuscript Figure 7e)
# -----------------------------------------------------------------------------
# The scheme as published is a two-node decision tree:
#
#   Node 1:  high ERBB2 SCNA  AND  high genome instability   -> S1
#   Node 2:  (not S1) AND high ADC trafficking AND low type 2 immunity -> S3
#   otherwise                                                -> S2
#
# Two things about it need to be reported honestly and are implemented below.
#
# First, the "high"/"low" cut-points were not pre-specified. They were chosen
# by sweeping a proportional cut-off from 0.2 to 0.8 in steps of 0.05 on this
# same cohort, with the outcome in view. A single reported p-value from the
# best cut-point of such a sweep is not a valid p-value: it is the minimum of
# many correlated tests. `sweep_cutpoints` runs the whole sweep and reports the
# full range of results, and `cutpoint_permutation_p` computes a selective-
# inference p-value by repeating the entire sweep under permuted treatment
# assignment, so the reported significance accounts for the search.
#
# Second, the groups were derived and evaluated in the same patients. No
# resampling can repair that; the code labels every output exploratory and the
# text must say the scheme requires prospective validation.
# -----------------------------------------------------------------------------

# Column names as they appear in the PREDIX dataset. Override with
# --s_group_spec if the manuscript used different operationalisations of
# "genome instability" (DNA_HRD and DNA_LOH_Del_burden are the alternatives)
# or of "type 2 immunity" (RNA_Mast-cells and RNA_TAM_M2 are the alternatives).
# Note the space in "RNA_Th2 cells" — it is not a hyphen.
S_GROUP_SPEC_DEFAULT = {
    "erbb2_scna":         "DNA_ERBB2_CNA",
    "genome_instability": "DNA_CNV_burden",
    "adc_trafficking":    "RNA_ADC_trafficking",
    "type2_immunity":     "RNA_Th2 cells",
}


def assign_s_groups(df, spec=None, q=0.5, q_map=None):
    """
    Assign S1 / S2 / S3 by the published two-node decision tree.

    Parameters
    ----------
    df    : dataframe holding the four driver variables.
    spec  : {role: column_name} mapping. Roles are 'erbb2_scna',
            'genome_instability', 'adc_trafficking', 'type2_immunity'.
    q     : shared proportional cut-off. A patient is "high" for a variable
            when its value is at or above the q-th quantile of that variable in
            this cohort, and "low" when at or below it.
    q_map : optional per-role override of q, for the sensitivity analysis.

    Returns
    -------
    labels    : pd.Series of 'S1' / 'S2' / 'S3', or None if columns are missing.
    cutpoints : {role: (column, quantile, threshold_value)}
    """
    spec = dict(S_GROUP_SPEC_DEFAULT if spec is None else spec)
    q_map = q_map or {}
    missing = [f"{role} -> {col}" for role, col in spec.items()
               if col not in df.columns]
    if missing:
        return None, {"missing": missing}

    cut = {}
    flag = {}
    for role, col in spec.items():
        qq = float(q_map.get(role, q))
        thr = float(df[col].quantile(qq))
        cut[role] = {"column": col, "quantile": qq, "threshold": thr}
        flag[role] = df[col].astype(float)

    high_erbb2  = flag["erbb2_scna"]         >= cut["erbb2_scna"]["threshold"]
    high_instab = flag["genome_instability"] >= cut["genome_instability"]["threshold"]
    high_adc    = flag["adc_trafficking"]    >= cut["adc_trafficking"]["threshold"]
    low_imm     = flag["type2_immunity"]     <= cut["type2_immunity"]["threshold"]

    s1 = (high_erbb2 & high_instab).fillna(False)
    s3 = ((~s1) & high_adc & low_imm).fillna(False)

    labels = pd.Series("S2", index=df.index, dtype=object)
    labels[s1] = "S1"
    labels[s3] = "S3"
    # Patients missing any driver cannot be classified; label them rather than
    # silently assigning them to S2, which would inflate the S2 denominator.
    any_missing = df[[spec[r] for r in spec]].isna().any(axis=1)
    labels[any_missing] = "Unclassifiable"
    return labels, cut


def adjusted_arm_or(df, group_col, arm_col, outcome_col, adjust_cols=("Clin_ER",)):
    """
    Per-group odds ratio for arm, from a logistic model adjusted for ER status.

    This mirrors the manuscript's primary biomarker model
    (logit(p) = b0 + b1*arm + b2*ER, fitted within each group) so the numbers
    are directly comparable with the published Figure 7f table, rather than
    being a different estimator that happens to be nearby.

    Falls back to the unadjusted cell-count odds ratio with a
    Haldane-Anscombe correction when the adjusted model cannot be fitted —
    which happens whenever a group x arm cell has zero or complete events.
    That fallback is flagged in the returned table, because an unbounded odds
    ratio is exactly the situation in which a confident-looking interval
    should not be trusted.
    """
    rows = []
    adjust_cols = [c for c in adjust_cols if c in df.columns]
    for grp in sorted(df[group_col].dropna().unique()):
        sub = df[df[group_col] == grp].dropna(
            subset=[arm_col, outcome_col] + adjust_cols)
        y = sub[outcome_col].astype(float).values
        a = sub[arm_col].astype(float).values
        n1, n0 = int((a == 1).sum()), int((a == 0).sum())
        e1 = int(np.nansum(y[a == 1])) if n1 else 0
        e0 = int(np.nansum(y[a == 0])) if n0 else 0

        or_val = lo = hi = np.nan
        method = "not estimable"
        separated = (n1 == 0 or n0 == 0
                     or min(e1, n1 - e1, e0, n0 - e0) == 0)

        if not separated and len(np.unique(y)) == 2:
            X = np.column_stack(
                [a] + [sub[c].astype(float).values for c in adjust_cols])
            try:
                _ll, beta, se = _fit_logit_manual(X, y)
                b_arm, se_arm = float(beta[1]), float(se[1])
                if np.isfinite(se_arm) and se_arm > 0:
                    or_val = float(np.exp(b_arm))
                    lo = float(np.exp(b_arm - 1.96 * se_arm))
                    hi = float(np.exp(b_arm + 1.96 * se_arm))
                    method = ("logistic, adjusted for "
                              + ", ".join(adjust_cols) if adjust_cols
                              else "logistic, unadjusted")
            except Exception:
                pass

        if not np.isfinite(or_val) and n1 and n0:
            corr = 0.5 if min(e1, n1 - e1, e0, n0 - e0) == 0 else 0.0
            odds1 = (e1 + corr) / (n1 - e1 + corr)
            odds0 = (e0 + corr) / (n0 - e0 + corr)
            if odds0 > 0:
                or_val = odds1 / odds0
                se_log = np.sqrt(1 / (e1 + corr) + 1 / (n1 - e1 + corr)
                                 + 1 / (e0 + corr) + 1 / (n0 - e0 + corr))
                lo = float(np.exp(np.log(or_val) - 1.96 * se_log))
                hi = float(np.exp(np.log(or_val) + 1.96 * se_log))
                method = ("unadjusted cell counts, Haldane-Anscombe corrected"
                          if corr else "unadjusted cell counts")

        rows.append({
            "group": grp,
            "n_DHP": n0, "events_DHP": e0,
            "pCR_DHP": (e0 / n0) if n0 else np.nan,
            "n_TDM1": n1, "events_TDM1": e1,
            "pCR_TDM1": (e1 / n1) if n1 else np.nan,
            "OR_TDM1_vs_DHP": or_val, "OR_ci_low": lo, "OR_ci_high": hi,
            "estimator": method,
            "sparse_or_separated_cell": separated,
        })
    return pd.DataFrame(rows)


def sweep_cutpoints(df, arm_col, outcome_col, spec=None,
                    q_grid=None, adjust_cols=("Clin_ER",)):
    """
    Repeat the S-group construction across the full cut-point grid.

    The manuscript's Methods state that continuous variables were dichotomised
    by "a parameter sweep from 0.2 to 0.8 in increments of 0.05". That sweep
    was optimised against the outcome on this cohort, so the reported result
    is the best of many correlated analyses. This function reports what the
    other analyses in that sweep produced.

    A scheme whose interaction p-value is small only at one or two cut-points
    out of thirteen is fragile; one that holds across most of the grid is a
    genuinely more robust finding. Either way the reader can now see which it
    is instead of only the selected value.
    """
    q_grid = np.round(np.arange(0.20, 0.8001, 0.05), 4) if q_grid is None else q_grid
    rows = []
    for q in q_grid:
        labels, cut = assign_s_groups(df, spec=spec, q=float(q))
        if labels is None:
            return pd.DataFrame(), cut
        d = df.copy()
        d["_group"] = labels
        d = d[d["_group"] != "Unclassifiable"]
        if d["_group"].nunique() < 2:
            continue
        inter = interaction_lrt_adjusted(d, "_group", arm_col, outcome_col,
                                         adjust_cols=adjust_cols)
        counts = d.groupby("_group").size().to_dict()
        rows.append({
            "cutpoint_quantile": float(q),
            "n_classified": int(len(d)),
            "n_S1": int(counts.get("S1", 0)),
            "n_S2": int(counts.get("S2", 0)),
            "n_S3": int(counts.get("S3", 0)),
            "interaction_LRT": inter["lrt_stat"],
            "interaction_df": inter["df"],
            "interaction_p": inter["p_value"],
            "n_events": inter["n_events"],
        })
    return pd.DataFrame(rows), None


def interaction_lrt_adjusted(df, group_col, arm_col, outcome_col,
                             adjust_cols=("Clin_ER",)):
    """
    Treatment-by-group interaction likelihood-ratio test, adjusted for ER.

    Nested models, matching the manuscript's Equation 2:
        M0: pCR ~ group + arm + ER
        M1: pCR ~ group + arm + ER + group:arm
    The LRT on the interaction block is the test. This is the quantity a
    differential-benefit claim rests on; separately elevated pCR rates within
    each arm are not evidence of differential benefit, because two groups can
    both respond well to their assigned treatment without the assignment
    having mattered.
    """
    adjust_cols = [c for c in adjust_cols if c in df.columns]
    need = [group_col, arm_col, outcome_col] + adjust_cols
    d = df[need].dropna().copy()
    out = {"lrt_stat": np.nan, "df": np.nan, "p_value": np.nan,
           "n": len(d), "n_events": int(np.nansum(d[outcome_col].values)),
           "adjusted_for": list(adjust_cols)}
    if len(d) < 20 or d[group_col].nunique() < 2 or d[arm_col].nunique() < 2:
        return out

    ref = d[group_col].astype(str).value_counts().idxmax()
    out["reference_group"] = ref
    dummies = pd.get_dummies(d[group_col].astype(str), prefix="g")
    dummies = dummies.drop(columns=[f"g_{ref}"], errors="ignore")
    G = dummies.values.astype(float)
    a = d[arm_col].astype(float).values.reshape(-1, 1)
    Z = (np.column_stack([d[c].astype(float).values for c in adjust_cols])
         if adjust_cols else np.empty((len(d), 0)))
    y = d[outcome_col].astype(float).values

    X0 = np.hstack([G, a, Z])
    X1 = np.hstack([G, a, Z, G * a])
    ll0 = _fit_logit_manual(X0, y)[0]
    ll1 = _fit_logit_manual(X1, y)[0]
    stat = 2.0 * (ll1 - ll0)
    dfree = X1.shape[1] - X0.shape[1]
    out["lrt_stat"] = float(stat)
    out["df"] = int(dfree)
    out["p_value"] = (float(stats.chi2.sf(max(stat, 0.0), dfree))
                      if dfree > 0 else np.nan)
    return out


def cutpoint_permutation_p(df, arm_col, outcome_col, spec=None,
                           q_grid=None, adjust_cols=("Clin_ER",),
                           n_perm=1000, seed=None):
    """
    Selective-inference p-value for the cut-point-optimised interaction.

    The observed statistic is the SMALLEST interaction p-value across the whole
    cut-point sweep — i.e. what the search actually reports.

    NULL HYPOTHESIS, stated precisely: permuting the (randomised) arm labels
    makes them exchangeable with respect to the outcome only under the SHARP
    null of no treatment effect of ANY form — the permutation destroys the
    arm main effect as well as the interaction. The test is therefore of "no
    treatment effect at all", which is STRONGER than the interaction LRT's
    null (which retains the arm main effect in M0). When a nonzero arm main
    effect exists, the reported permutation p is approximate for the
    interaction-only question; in PREDIX HER2 the arm difference in pCR is
    small, so the approximation is mild. An exact alternative preserving the
    main effect is a parametric bootstrap under M0 (fit pCR ~ group + arm +
    ER, simulate y*, re-run the sweep per simulation).

    The corrected p-value is the fraction of permutations whose minimum sweep
    p-value is at least as extreme as the observed one. It answers the question
    a nominal p-value cannot: how often would a sweep this wide produce a
    result this strong when nothing is there?

    seed: None resolves to the module-global BOOT_SEED at call time so the
    --seed CLI override applies.

    Returns dict with observed min p, the cut-point that achieved it, the
    permutation p-value, and the number of permutations.
    """
    if seed is None: seed = BOOT_SEED
    rng = np.random.default_rng(seed)
    obs, _ = sweep_cutpoints(df, arm_col, outcome_col, spec=spec,
                             q_grid=q_grid, adjust_cols=adjust_cols)
    out = {"observed_min_p": np.nan, "best_cutpoint": np.nan,
           "permutation_p": np.nan, "n_perm": n_perm,
           "n_cutpoints_searched": 0 if obs.empty else int(len(obs))}
    if obs.empty or obs["interaction_p"].isna().all():
        return out

    obs_min = float(obs["interaction_p"].min())
    out["observed_min_p"] = obs_min
    out["best_cutpoint"] = float(
        obs.loc[obs["interaction_p"].idxmin(), "cutpoint_quantile"])

    d = df.copy()
    n_extreme = 0
    n_valid = 0
    for _ in range(n_perm):
        d[arm_col] = rng.permutation(df[arm_col].values)
        perm, _ = sweep_cutpoints(d, arm_col, outcome_col, spec=spec,
                                  q_grid=q_grid, adjust_cols=adjust_cols)
        if perm.empty or perm["interaction_p"].isna().all():
            continue
        n_valid += 1
        if float(perm["interaction_p"].min()) <= obs_min:
            n_extreme += 1

    if n_valid:
        # Add-one correction so the p-value can never be exactly zero, which a
        # finite permutation set cannot support.
        out["permutation_p"] = float((n_extreme + 1) / (n_valid + 1))
        out["n_perm"] = n_valid
    return out


# =============================================================================
# SECTION 7 — EXPLORATORY SURVIVAL ANALYSIS
# =============================================================================
# The recurrence analysis rests on a small number of events. It is reported as
# exploratory and hypothesis-generating: event counts are stated everywhere,
# p-values are Benjamini-Hochberg corrected across the tested variables, and
# no inferential claim is drawn from it.
# =============================================================================

def kaplan_meier(time, event):
    """
    Kaplan-Meier estimator. Returns (times, survival, at_risk, n_events).

    Implemented directly rather than via lifelines so the repository has no
    dependency a reviewer cannot install from a bare environment.
    """
    time = np.asarray(time, dtype=float)
    event = np.asarray(event, dtype=float)
    ok = np.isfinite(time) & np.isfinite(event)
    time, event = time[ok], event[ok]
    if len(time) == 0:
        return np.array([]), np.array([]), np.array([]), np.array([])

    order = np.argsort(time)
    time, event = time[order], event[order]
    uniq = np.unique(time[event == 1])

    surv, at_risk_out, d_out = [], [], []
    s = 1.0
    for t in uniq:
        n_at_risk = int(np.sum(time >= t))
        d = int(np.sum((time == t) & (event == 1)))
        if n_at_risk > 0:
            s *= (1.0 - d / n_at_risk)
        surv.append(s)
        at_risk_out.append(n_at_risk)
        d_out.append(d)
    return uniq, np.array(surv), np.array(at_risk_out), np.array(d_out)


def logrank_test(time, event, group):
    """
    Two-sample log-rank test (Mantel-Cox).

    Returns dict with chi2, p_value, observed and expected events per group.
    """
    time = np.asarray(time, dtype=float)
    event = np.asarray(event, dtype=float)
    group = np.asarray(group)
    ok = np.isfinite(time) & np.isfinite(event)
    time, event, group = time[ok], event[ok], group[ok]

    levels = np.unique(group)
    out = {"chi2": np.nan, "p_value": np.nan, "levels": list(levels),
           "observed": {}, "expected": {},
           "n": len(time), "n_events": int(event.sum())}
    if len(levels) != 2 or len(time) < 5 or event.sum() < 2:
        return out

    g1 = group == levels[0]
    event_times = np.unique(time[event == 1])
    o1 = e1 = v = 0.0
    for t in event_times:
        n_risk = float(np.sum(time >= t))
        n_risk1 = float(np.sum((time >= t) & g1))
        d = float(np.sum((time == t) & (event == 1)))
        d1 = float(np.sum((time == t) & (event == 1) & g1))
        if n_risk <= 1:
            continue
        exp1 = d * n_risk1 / n_risk
        var = (d * (n_risk1 / n_risk) * (1 - n_risk1 / n_risk)
               * (n_risk - d) / (n_risk - 1))
        o1 += d1
        e1 += exp1
        v += var

    out["observed"] = {str(levels[0]): float(o1),
                       str(levels[1]): float(event.sum() - o1)}
    out["expected"] = {str(levels[0]): float(e1),
                       str(levels[1]): float(event.sum() - e1)}
    if v <= 0:
        return out
    chi2 = (o1 - e1) ** 2 / v
    out["chi2"] = float(chi2)
    out["p_value"] = float(stats.chi2.sf(chi2, 1))
    return out


def cox_model(time, event, X, col_names=None):
    """
    Cox proportional-hazards model with Breslow tie handling, univariable or
    multivariable.

    Maximises the Breslow partial likelihood with scipy. Returns per-covariate
    hazard ratios, 95% Wald intervals and Wald p-values; standard errors come
    from the numerical Hessian of the negative log partial likelihood.

    Covariates are standardised internally for numerical conditioning and the
    coefficients are rescaled to original units before being returned.

    An events-per-variable value is returned on every fit. With 17 recurrence
    events, a model adjusting for treatment arm, tumour size, nodal status and
    ER status has an EPV of 17/5 = 3.4 — far below the conventional minimum of
    10, and low enough that Wald intervals are unreliable and the model can
    approach separation. The caller is expected to report this number rather
    than the hazard ratio alone.
    """
    time = np.asarray(time, dtype=float)
    event = np.asarray(event, dtype=float)
    X = np.atleast_2d(np.asarray(X, dtype=float))
    if X.shape[0] != len(time):
        X = X.T
    n_cov = X.shape[1]
    col_names = list(col_names) if col_names is not None else [
        f"x{i}" for i in range(n_cov)]

    ok = (np.isfinite(time) & np.isfinite(event)
          & np.isfinite(X).all(axis=1))
    time, event, X = time[ok], event[ok], X[ok]

    n_events = int(event.sum())
    out = {"n": len(time), "n_events": n_events,
           "epv": (n_events / n_cov) if n_cov else np.nan,
           "n_covariates": n_cov, "converged": False,
           "terms": pd.DataFrame()}
    if len(time) < 5 or n_events < 2 or n_cov == 0:
        return out
    sd = X.std(axis=0)
    if np.any(sd == 0):
        return out

    mu = X.mean(axis=0)
    Xs = (X - mu) / sd

    order = np.argsort(-time)          # descending time -> cumulative risk set
    t_s, e_s, X_s = time[order], event[order], Xs[order]
    event_pos = np.where(e_s == 1)[0]

    # For Breslow ties, every subject dying at time t sees the risk set of all
    # subjects with time >= t. Precompute, for each event, the last index whose
    # time equals that event's time.
    last_tied = np.empty(len(event_pos), dtype=int)
    for k, i in enumerate(event_pos):
        j = i
        while j + 1 < len(t_s) and t_s[j + 1] == t_s[i]:
            j += 1
        last_tied[k] = j

    def neg_log_pl(beta):
        eta = np.clip(X_s @ beta, -50, 50)
        cum = np.cumsum(np.exp(eta))
        return -float(np.sum(eta[event_pos] - np.log(cum[last_tied])))

    try:
        res = minimize(neg_log_pl, x0=np.zeros(n_cov), method="BFGS")
        beta = np.asarray(res.x, dtype=float)
        out["converged"] = bool(res.success)
    except Exception:
        return out

    # Numerical Hessian of the negative log partial likelihood.
    h = 1e-4
    H = np.zeros((n_cov, n_cov))
    f0 = neg_log_pl(beta)
    for i in range(n_cov):
        for j in range(i, n_cov):
            ei = np.zeros(n_cov); ei[i] = h
            ej = np.zeros(n_cov); ej[j] = h
            fpp = neg_log_pl(beta + ei + ej)
            fpm = neg_log_pl(beta + ei - ej)
            fmp = neg_log_pl(beta - ei + ej)
            fmm = neg_log_pl(beta - ei - ej)
            H[i, j] = H[j, i] = (fpp - fpm - fmp + fmm) / (4 * h * h)
    try:
        cov = np.linalg.inv(H)
        se_s = np.sqrt(np.clip(np.diag(cov), 0, np.inf))
    except np.linalg.LinAlgError:
        se_s = np.full(n_cov, np.nan)

    coef = beta / sd
    se = se_s / sd
    rows = []
    for i, name in enumerate(col_names):
        z = coef[i] / se[i] if np.isfinite(se[i]) and se[i] > 0 else np.nan
        rows.append({
            "term": name,
            "coef": float(coef[i]),
            "hazard_ratio": float(np.exp(coef[i])),
            "se": float(se[i]) if np.isfinite(se[i]) else np.nan,
            "HR_ci_low": (float(np.exp(coef[i] - 1.96 * se[i]))
                          if np.isfinite(se[i]) else np.nan),
            "HR_ci_high": (float(np.exp(coef[i] + 1.96 * se[i]))
                           if np.isfinite(se[i]) else np.nan),
            "z": float(z) if np.isfinite(z) else np.nan,
            "p_value": (float(2 * stats.norm.sf(abs(z)))
                        if np.isfinite(z) else np.nan),
        })
    out["terms"] = pd.DataFrame(rows)
    return out


def cox_univariable(time, event, x):
    """Convenience wrapper: single-covariate Cox model, flat return dict."""
    res = cox_model(time, event, np.asarray(x, float).reshape(-1, 1), ["x"])
    out = {"coef": np.nan, "hr": np.nan, "ci_low": np.nan, "ci_high": np.nan,
           "se": np.nan, "z": np.nan, "p_value": np.nan,
           "n": res["n"], "n_events": res["n_events"], "epv": res["epv"]}
    if not res["terms"].empty:
        t = res["terms"].iloc[0]
        out.update({"coef": t["coef"], "hr": t["hazard_ratio"],
                    "ci_low": t["HR_ci_low"], "ci_high": t["HR_ci_high"],
                    "se": t["se"], "z": t["z"], "p_value": t["p_value"]})
    return out


def exploratory_survival(df, time_col, event_col, variables,
                         adjust_cols=()):
    """
    Exploratory recurrence analysis over a set of candidate variables.

    Each variable gets a Cox model; binary variables additionally get a
    log-rank test. When `adjust_cols` is supplied the Cox model is
    multivariable, matching the manuscript's adjustment set (treatment arm,
    tumour size, nodal status, ER status), and the realised events-per-variable
    is reported on every row.

    p-values are Benjamini-Hochberg corrected ACROSS the variables tested. The
    original analysis reported these at nominal p, which is the multiplicity
    limitation now acknowledged in the text.
    """
    rows = []
    n_events_total = int(np.nansum(df[event_col].values))
    adjust_cols = [c for c in adjust_cols if c in df.columns]

    for var in variables:
        if var not in df.columns or var in adjust_cols:
            continue
        cols = [time_col, event_col, var] + adjust_cols
        sub = df[cols].dropna()
        if len(sub) < 10 or sub[var].nunique() < 2:
            continue

        design = [sub[var].astype(float).values] + [
            sub[c].astype(float).values for c in adjust_cols]
        names = [var] + list(adjust_cols)
        res = cox_model(sub[time_col].values, sub[event_col].values,
                        np.column_stack(design), names)

        if res["terms"].empty:
            continue
        t = res["terms"].iloc[0]
        row = {"variable": var, "n": res["n"], "n_events": res["n_events"],
               "n_covariates": res["n_covariates"],
               "events_per_variable": res["epv"],
               "adjusted_for": ", ".join(adjust_cols) or "nothing (univariable)",
               "hazard_ratio": t["hazard_ratio"],
               "HR_ci_low": t["HR_ci_low"], "HR_ci_high": t["HR_ci_high"],
               "p_cox": t["p_value"], "p_logrank": np.nan,
               "converged": res["converged"]}
        if sub[var].nunique() == 2:
            lr = logrank_test(sub[time_col].values, sub[event_col].values,
                              sub[var].values)
            row["p_logrank"] = lr["p_value"]
        # Flag estimates that the data cannot really support: an enormous or
        # vanishing hazard ratio at this event count signals near-separation,
        # not a strong effect.
        hr = t["hazard_ratio"]
        row["unstable_estimate"] = bool(
            (np.isfinite(hr) and (hr > 20 or hr < 0.05))
            or not np.isfinite(t["se"]) or res["epv"] < 5)
        rows.append(row)

    df_out = pd.DataFrame(rows)
    if not df_out.empty:
        df_out["q_cox_BH"] = benjamini_hochberg(df_out["p_cox"].values)
        if df_out["p_logrank"].notna().any():
            df_out["q_logrank_BH"] = benjamini_hochberg(
                df_out["p_logrank"].values)
        df_out["total_events_in_cohort"] = n_events_total
        df_out["interpretation"] = (
            "EXPLORATORY — hypothesis-generating only; "
            f"{n_events_total} events limit power severely")
        df_out = df_out.sort_values("p_cox").reset_index(drop=True)
    return df_out


# =============================================================================
# SECTION 8 — EXCEL AND FIGURE WRITERS
# =============================================================================

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_NOTE_FONT = Font(italic=True, size=9, color="555555")


def _write_sheet(wb, title, df, note=None, first=False):
    """Write a DataFrame to a styled worksheet; returns the worksheet."""
    ws = wb.active if first else wb.create_sheet()
    ws.title = title[:31]
    r = 1
    if note:
        for line in str(note).split("\n"):
            ws.cell(row=r, column=1, value=line).font = _NOTE_FONT
            r += 1
        r += 1
    if df is None or df.empty:
        ws.cell(row=r, column=1, value="No data available for this analysis.")
        return ws
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=r, column=c, value=str(col))
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, (_, row) in enumerate(df.iterrows(), start=r + 1):
        for c, col in enumerate(df.columns, start=1):
            v = row[col]
            if isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = None if not np.isfinite(v) else round(float(v), 6)
            elif isinstance(v, (np.bool_,)):
                v = bool(v)
            elif isinstance(v, (list, tuple, dict)):
                v = str(v)
            ws.cell(row=i, column=c, value=v)
    for c, col in enumerate(df.columns, start=1):
        # pandas 3: astype(str) keeps NaN as NaN (str dtype is NaN-aware), so
        # an all-NaN column makes .str.len().max() NaN and int(NaN) raises.
        mx = df[col].astype(str).str.len().max() if len(df) else np.nan
        width = max(12, min(46, int(mx) + 3)) if pd.notna(mx) else 14
        ws.column_dimensions[get_column_letter(c)].width = max(width, len(str(col)) + 3)
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    return ws


def _savefig(fig, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path)
    plt.close(fig)
    print(f"  -> {path.name}")


def _asym_err(centre, low, high):
    """
    Build the [lower, upper] error-bar arrays matplotlib expects, clipped at 0.

    A Wilson interval always contains its point estimate in exact arithmetic,
    but when the observed proportion is exactly 0 or 1 the bound can land a
    few ulps on the wrong side of it. matplotlib rejects the resulting -1e-16
    outright with "'yerr' must not contain negative values", which kills the
    whole figure. Clipping here is the correct fix: the quantity is a distance
    and cannot be negative.
    """
    centre = np.asarray(centre, dtype=float)
    low    = np.asarray(low, dtype=float)
    high   = np.asarray(high, dtype=float)
    lower = np.clip(centre - low, 0.0, None)
    upper = np.clip(high - centre, 0.0, None)
    return [np.nan_to_num(lower), np.nan_to_num(upper)]


# =============================================================================
# SECTION 9 — DRIVERS (one per reviewer point)
# =============================================================================

def run_performance_ci(results, td):
    """
    Performance of every model (discovery and consensus): metric per CV
    repeat averaged over repeats, with patient-level cluster-bootstrap CIs.
    """
    import zlib
    rows = []
    for sc, entry in results.items():
        for source, blob in (("consensus", entry["consensus"]),
                             ("discovery", entry["discovery"])):
            if blob is None:
                continue
            for mod in ALL_MODELS:
                try:
                    rm = _repeat_matrix(blob, source, mod)
                except (KeyError, ValueError) as e:
                    print(f"  [CI] {sc}/{source}/{mod}: {e}")
                    continue
                if rm is None or rm.n_patients == 0:
                    continue
                if rm.incomplete_repeats:
                    print(f"  [CI] {sc}/{source}/{mod}: {rm.incomplete_repeats} "
                          f"incomplete repeat(s) — metrics use available entries.")
                row = {"scenario": sc, "source": source, "model": mod,
                       "n_patients": rm.n_patients, "n_events": rm.n_events,
                       "n_cv_repeats": rm.n_repeats,
                       "n_outer_folds": rm.n_folds}
                for metric in ("AUROC", "AUPRC", "Brier"):
                    # zlib.crc32, not hash(): Python string hashing is salted
                    # per process, so hash()-derived seeds change on every
                    # run and the CI endpoints wobble despite the pipeline
                    # advertising seeded reproducibility.
                    tag = zlib.crc32(f"{sc}|{source}|{mod}|{metric}".encode())
                    res = CE.bootstrap_repeat_metric_ci(
                        rm.P, rm.y, metric=metric, n_boot=N_BOOT,
                        seed=BOOT_SEED + tag % 10000)
                    row[f"{metric}"] = res["estimate"]
                    row[f"{metric}_CI_low"] = res["ci_low"]
                    row[f"{metric}_CI_high"] = res["ci_high"]
                    row[f"{metric}_formatted"] = format_ci(res)
                # Continuity with the submitted manuscript, whose point
                # estimates were means of per-fold AUROC. Reported without an
                # interval: it is the same quantity up to between-fold
                # calibration heterogeneity, and its "±" SD was never a CI.
                if source == "consensus":
                    getter = ((lambda f: f["fused_y_pred"]) if mod == FUSED
                              else (lambda f, m=mod: f["unimodal_y_pred"][m]))
                    row["AUROC_mean_per_fold_(original_estimand,_no_CI)"] = \
                        CE.mean_fold_metric(blob["folds"], getter, "AUROC")
                else:
                    row["AUROC_mean_per_fold_(original_estimand,_no_CI)"] = \
                        CE.mean_fold_metric(blob.get(mod, []),
                                            lambda f: f["y_pred"], "AUROC")
                rows.append(row)

    df = pd.DataFrame(rows)
    wb = openpyxl.Workbook()
    _write_sheet(
        wb, "Performance_patient_CI", df, first=True,
        note=("CROSS-VALIDATED PERFORMANCE WITH PATIENT-LEVEL CLUSTER-BOOTSTRAP "
              "95% CONFIDENCE INTERVALS.\n"
              "Estimand: in each repeat of stratified K-fold cross-validation every "
              "patient receives exactly one out-of-fold prediction; the metric is "
              "computed on that complete out-of-fold vector and averaged over the "
              "n_cv_repeats repeats (200 global, 100 per arm). This is a single-model "
              "quantity — predictions are never averaged across repeats or models.\n"
              f"Uncertainty: {N_BOOT} stratified patient-level resamples (observed pCR "
              "count preserved). A resampled patient carries ALL its repeat "
              "predictions, so the interval reflects the number of independent "
              "patients, not the number of (patient, fold) rows.\n"
              "The '±' values in the original submission were the standard deviation "
              "of per-fold AUROC across repeated cross-validation and are NOT "
              "confidence intervals; the corresponding point estimate is kept in the "
              "last column for continuity only.\n"
              "Discovery = per-fold winner signature/classifier (fully nested, no "
              "selection optimism). Consensus = frozen consensus signature and "
              "classifier re-evaluated on the same splits (fold-refit weights and "
              "fusion; carries the optimism of having chosen the signature with "
              "knowledge of all outcomes)."))
    path = td / "revision_performance_CI.xlsx"
    wb.save(path)
    print(f"  -> {path.name}")
    return df


def run_calibration(results, td, fd):
    """Reliability curves, slope/intercept and Brier with CI (per CV repeat)."""
    import zlib
    summary_rows = []
    reliability_frames = []

    for sc, entry in results.items():
        blob = entry["consensus"]
        source = "consensus"
        if blob is None:
            blob = entry["discovery"]
            source = "discovery"
        if blob is None:
            continue
        try:
            rm = _repeat_matrix(blob, source, FUSED)
        except (KeyError, ValueError) as e:
            print(f"  [CAL] {sc}: {e}")
            continue
        if rm is None or rm.n_patients < 10:
            continue

        tag = zlib.crc32(f"cal|{sc}|{source}".encode()) % 10000
        cal = CE.calibration_repeat_summary(rm.P, rm.y, n_boot=N_BOOT,
                                            seed=BOOT_SEED + tag,
                                            n_bins=N_CAL_BINS)
        summary_rows.append({
            "scenario": sc, "source": source, "model": FUSED,
            "n": cal["n"], "n_events": cal["n_events"],
            "n_cv_repeats": cal["n_repeats"],
            "observed_pCR_rate": cal["observed_rate"],
            "mean_predicted": cal["mean_predicted"],
            "calibration_slope": cal["slope"],
            "slope_CI_low": cal["slope_ci"][0],
            "slope_CI_high": cal["slope_ci"][1],
            "calibration_intercept": cal["intercept"],
            "intercept_CI_low": cal["intercept_ci"][0],
            "intercept_CI_high": cal["intercept_ci"][1],
            "brier": cal["brier"],
            "brier_CI_low": cal["brier_ci"][0],
            "brier_CI_high": cal["brier_ci"][1],
            "ECE": cal["ece"],
        })
        rel = CE.reliability_pooled(rm.P, rm.y, n_bins=N_CAL_BINS,
                                    n_boot=N_BOOT, seed=BOOT_SEED + tag + 1)
        if not rel.empty:
            rel.insert(0, "scenario", sc)
            reliability_frames.append(rel)

    df_sum = pd.DataFrame(summary_rows)
    df_rel = (pd.concat(reliability_frames, ignore_index=True)
              if reliability_frames else pd.DataFrame())

    wb = openpyxl.Workbook()
    _write_sheet(
        wb, "Calibration_summary", df_sum, first=True,
        note=("CALIBRATION OF THE CONSENSUS FUSED MODEL on out-of-fold predictions.\n"
              "Slope and intercept from logit(P(pCR)) = intercept + slope * "
              "logit(p_predicted), fitted on each cross-validation repeat's complete "
              "out-of-fold vector and averaged over repeats (Brier and ECE likewise). "
              "Slope 1 and intercept 0 indicate perfect calibration; slope < 1 means "
              "predictions are too extreme, slope > 1 too compressed.\n"
              "Probabilities are Platt-recalibrated WITHIN the cross-validation, so a "
              "slope near 1 shows that recalibration transferred to held-out patients.\n"
              f"Confidence intervals: {N_BOOT} stratified patient-level cluster-"
              "bootstrap resamples (a resampled patient carries all its repeat "
              "predictions). Predictions are never averaged across repeats: doing "
              "so scores an ensemble whose compressed probabilities inflate the "
              "slope."))
    _write_sheet(
        wb, "Reliability_bins", df_rel,
        note=("Reliability-curve data from ALL (repeat, patient) out-of-fold "
              "predictions pooled: equal-count (quantile) bins on predicted risk; "
              "equal-width bins leave several bins empty at this sample size.\n"
              "n_rows counts (patient, repeat) predictions; n_patients_distinct is the "
              "number of different patients contributing to the bin. obs_ci is a "
              "patient-level cluster-bootstrap interval on the observed fraction "
              "(bins re-derived per resample), so it respects that each patient "
              "contributes many correlated rows."))
    path = td / "revision_calibration.xlsx"
    wb.save(path)
    print(f"  -> {path.name}")

    # ── Figure ───────────────────────────────────────────────────────────────
    if not df_rel.empty:
        scs = [s for s in SCENARIOS if s in set(df_rel["scenario"])]
        fig, axes = plt.subplots(1, len(scs), figsize=(4.6 * len(scs), 4.6),
                                 squeeze=False)
        for ax, sc in zip(axes[0], scs):
            sub = df_rel[df_rel["scenario"] == sc]
            row = df_sum[df_sum["scenario"] == sc]
            ax.plot([0, 1], [0, 1], ls=":", c="#999", lw=1.1,
                    label="perfect calibration")
            ax.errorbar(sub["mean_predicted"], sub["observed"],
                        yerr=_asym_err(sub["observed"], sub["obs_ci_low"],
                                       sub["obs_ci_high"]),
                        fmt="o-", color=SC_COL.get(sc, "#333"),
                        ms=5, lw=1.4, capsize=3, label="observed")
            if not row.empty:
                r = row.iloc[0]
                ax.text(0.03, 0.97,
                        f"slope {r['calibration_slope']:.2f} "
                        f"[{r['slope_CI_low']:.2f}–{r['slope_CI_high']:.2f}]\n"
                        f"intercept {r['calibration_intercept']:.2f} "
                        f"[{r['intercept_CI_low']:.2f}–{r['intercept_CI_high']:.2f}]\n"
                        f"Brier {r['brier']:.3f} "
                        f"[{r['brier_CI_low']:.3f}–{r['brier_CI_high']:.3f}]\n"
                        f"n={int(r['n'])}, {int(r['n_events'])} events",
                        transform=ax.transAxes, va="top", ha="left", fontsize=7.5,
                        bbox=dict(boxstyle="round,pad=0.35", fc="white",
                                  ec="#cccccc", alpha=0.92))
            ax.set_xlim(0, 1); ax.set_ylim(0, 1)
            ax.set_xlabel("Mean predicted probability of pCR")
            ax.set_ylabel("Observed pCR fraction")
            ax.set_title(sc, color=SC_COL.get(sc, "#333"))
            ax.grid(alpha=0.3)
            ax.legend(loc="lower right", fontsize=7)
        fig.suptitle(
            "Calibration of the consensus multimodal model\n"
            "All out-of-fold predictions pooled (equal-count bins, patient-level "
            "cluster-bootstrap intervals); slope/intercept/Brier per CV repeat, "
            "averaged",
            fontsize=10.5, fontweight="bold", y=1.03)
        plt.tight_layout()
        _savefig(fig, fd / "revfig01_calibration.pdf")

    return df_sum, df_rel


def run_stability(results, td, fd):
    """Feature-selection frequency and modality-weight stability."""
    sel_frames, weight_frames = [], []
    raw_weights = {}   # {scenario: {modality: [per-fold fusion weights]}}
    for sc, entry in results.items():
        disc = entry["discovery"]
        if disc is None:
            continue
        thresh = STABILITY_THRESH.get(sc, 0.5)
        for mod in UNIMODALS:
            folds = disc.get(mod, [])
            if not folds:
                continue
            d = selection_frequency(folds, thresh)
            if d.empty:
                continue
            d.insert(0, "modality", mod)
            d.insert(0, "scenario", sc)
            sel_frames.append(d)
        fus = disc.get(FUSED, [])
        if fus:
            w = modality_weight_stability(fus)
            if not w.empty:
                w.insert(0, "scenario", sc)
                weight_frames.append(w)
            # Raw per-fold weights, for the distribution figure below.
            rw = defaultdict(list)
            for f in fus:
                for mod, val in f.get("modality_weights", {}).items():
                    rw[mod].append(float(val))
            raw_weights[sc] = rw

    df_sel = (pd.concat(sel_frames, ignore_index=True)
              if sel_frames else pd.DataFrame())
    df_w = (pd.concat(weight_frames, ignore_index=True)
            if weight_frames else pd.DataFrame())

    wb = openpyxl.Workbook()
    _write_sheet(
        wb, "Feature_selection_stability", df_sel, first=True,
        note=("FEATURE-SELECTION STABILITY ACROSS OUTER CROSS-VALIDATION FOLDS.\n"
              "selection_freq uses all outer folds as the denominator. "
              "selection_freq_eligible uses only folds in which the feature survived "
              "that fold's near-zero-variance/correlation/univariate-screen filters "
              "and could therefore have been selected (the fold's recorded candidate "
              "pool) — a feature culled before selection never had the opportunity, "
              "and scoring it as 'not selected' understates its stability. If the "
              "PKL predates candidate-pool recording, the all-folds denominator is "
              "used for both columns (conservative).\n"
              "wilson_low / wilson_high are Wilson score intervals on the eligible "
              "frequency.\n"
              "'stable' applies the PRE-SPECIFIED threshold (Global 0.60, arms 0.50). "
              "'stable_ci_supported' is the stricter criterion that the Wilson lower "
              "bound also clears the threshold. Features should be described as robust "
              "determinants only when they satisfy these criteria."))
    _write_sheet(
        wb, "Modality_weight_stability", df_w,
        note=("STABILITY OF THE LATE-FUSION MODALITY WEIGHTS.\n"
              "The fusion layer is an elastic-net (L1+L2) logistic model over the five "
              "calibrated modality probability streams. The L1 penalty sets "
              "non-contributing modalities to exactly zero.\n"
              "selection_rate is the fraction of folds in which the modality received a "
              "non-zero weight. sign_consistency is the fraction of those folds in which "
              "the weight took the dominant sign — a mean weight near zero with low sign "
              "consistency means the sign flips across folds, which a mean and SD alone "
              "would hide."))
    path = td / "revision_stability.xlsx"
    wb.save(path)
    print(f"  -> {path.name}")

    # ── Figure: selection frequency per modality x scenario ──────────────────
    if not df_sel.empty:
        scs = [s for s in SCENARIOS if s in set(df_sel["scenario"])]
        mods = [m for m in UNIMODALS if m in set(df_sel["modality"])]
        fig, axes = plt.subplots(len(mods), len(scs),
                                 figsize=(5.0 * len(scs), 2.9 * len(mods)),
                                 squeeze=False)
        for ri, mod in enumerate(mods):
            for ci, sc in enumerate(scs):
                ax = axes[ri][ci]
                sub = (df_sel[(df_sel["scenario"] == sc)
                              & (df_sel["modality"] == mod)]
                       .head(12).iloc[::-1])
                if sub.empty:
                    ax.text(0.5, 0.5, "no data", ha="center", va="center",
                            transform=ax.transAxes, color="#888", fontsize=8)
                    ax.set_xticks([]); ax.set_yticks([])
                    continue
                ypos = np.arange(len(sub))
                thr = STABILITY_THRESH.get(sc, 0.5)
                colors = [MOD_COLOR[mod] if v >= thr else "#c9c9c9"
                          for v in sub["selection_freq_eligible"]]
                ax.barh(ypos, sub["selection_freq_eligible"], color=colors,
                        height=0.68, edgecolor="white")
                ax.errorbar(sub["selection_freq_eligible"], ypos,
                            xerr=_asym_err(sub["selection_freq_eligible"],
                                           sub["wilson_low"],
                                           sub["wilson_high"]),
                            fmt="none", ecolor="#444", elinewidth=0.8, capsize=2)
                ax.axvline(thr, color="#c62828", ls="--", lw=1.0)
                ax.set_yticks(ypos)
                ax.set_yticklabels(sub["feature"], fontsize=6.8)
                ax.set_xlim(0, 1.02)
                if ri == len(mods) - 1:
                    ax.set_xlabel("selection frequency (eligible folds)", fontsize=8)
                if ri == 0:
                    ax.set_title(sc, color=SC_COL.get(sc, "#333"), fontsize=10)
                if ci == 0:
                    ax.set_ylabel(mod, fontsize=9, fontweight="bold",
                                  color=MOD_COLOR[mod])
                ax.grid(axis="x", alpha=0.3)
        fig.suptitle(
            "Feature-selection stability across cross-validation folds\n"
            "Wilson 95% intervals; dashed line = pre-specified stability threshold; "
            "grey bars fall below it",
            fontsize=11, fontweight="bold", y=1.005)
        plt.tight_layout()
        _savefig(fig, fd / "revfig02_selection_stability.pdf")

    # ── Figure: fold-wise distribution of the late-fusion modality weights ───
    # Reviewer #5 ("no coherent integrative analysis"): the fusion layer's
    # interpretable output is the per-modality weight, and its credibility
    # rests on the weights being stable across folds. One panel per scenario:
    # the full per-fold weight distribution per modality, with the selection
    # rate (fraction of folds with a non-zero weight, Wilson 95% CI) printed
    # alongside.
    scs_w = [s for s in SCENARIOS if s in raw_weights and raw_weights[s]]
    if scs_w and not df_w.empty:
        fig, axes = plt.subplots(1, len(scs_w),
                                 figsize=(4.6 * len(scs_w), 3.4),
                                 squeeze=False)
        rng = np.random.default_rng(7)   # jitter only; cosmetic
        for ci, sc in enumerate(scs_w):
            ax = axes[0][ci]
            rw = raw_weights[sc]
            mods = [m for m in MOD_COLOR if m in rw] or sorted(rw)
            sub_w = df_w[df_w["scenario"] == sc].set_index("modality")
            for yi, mod in enumerate(mods):
                vals = np.asarray(rw[mod], dtype=float)
                jitter = (rng.random(len(vals)) - 0.5) * 0.30
                ax.scatter(vals, np.full(len(vals), yi) + jitter, s=9,
                           color=MOD_COLOR.get(mod, "#666"), alpha=0.45,
                           linewidths=0)
                ax.plot([np.median(vals)] * 2, [yi - 0.28, yi + 0.28],
                        color=MOD_COLOR.get(mod, "#666"), lw=2.2)
                if mod in sub_w.index:
                    r = sub_w.loc[mod]
                    ax.text(1.02, yi,
                            f"{r['selection_rate']*100:.0f}% "
                            f"[{r['selection_rate_ci_low']*100:.0f}"
                            f"–{r['selection_rate_ci_high']*100:.0f}]",
                            transform=ax.get_yaxis_transform(),
                            fontsize=6.8, va="center", color="#444")
            ax.axvline(0, color="#999", lw=0.8, ls=":")
            ax.set_yticks(range(len(mods)))
            ax.set_yticklabels(mods, fontsize=8)
            ax.invert_yaxis()
            ax.set_title(sc, color=SC_COL.get(sc, "#333"), fontsize=10)
            ax.set_xlabel("fusion weight (per outer fold)", fontsize=8)
            ax.grid(axis="x", alpha=0.3)
        fig.suptitle(
            "Late-fusion modality weights across cross-validation folds\n"
            "dots = per-fold elastic-net coefficients; bar = median; "
            "right margin = selection rate (non-zero weight) with Wilson 95% CI",
            fontsize=10.5, fontweight="bold", y=1.02)
        plt.tight_layout()
        _savefig(fig, fd / "revfig08_fusion_weights.pdf")

    return df_sel, df_w


def run_epv(results, td, fd):
    """Per-fold event counts and realised events-per-variable."""
    frames = []
    for sc, entry in results.items():
        disc = entry["discovery"]
        if disc is None:
            continue
        for mod in ALL_MODELS:
            folds = disc.get(mod, [])
            if folds:
                frames.append(epv_table(folds, sc, mod))
    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    df_sum = summarise_epv(df)

    wb = openpyxl.Workbook()
    _write_sheet(
        wb, "EPV_summary", df_sum, first=True,
        note=("REALISED EVENTS-PER-VARIABLE, summarised per scenario and model.\n"
              "The pipeline caps signature size at EPV = 5 during discovery. That cap "
              "is a design constraint; the realised EPV differs from it because of the "
              "minimum-signature-size floor and because features can be dropped by a "
              "fold's preprocessing after the cap was applied.\n"
              "pct_folds_epv_below_5 is the fraction of folds where the realised EPV "
              "fell below the intended cap — the folds most exposed to overfitting."))
    _write_sheet(
        wb, "EPV_per_fold", df,
        note=("PER-FOLD pCR EVENT COUNTS AND REALISED EVENTS-PER-VARIABLE.\n"
              "n_events_test is the number of pCR events in that fold's held-out test "
              "set. n_events_train_expanded is the event count in the training set the "
              "signature was derived from.\n"
              "n_candidates_after_screen is the number of features that survived that "
              "fold's preprocessing INCLUDING the in-fold univariate screen — i.e. the "
              "size of the candidate pool the model actually chose from in that fold."))
    path = td / "revision_epv_per_fold.xlsx"
    wb.save(path)
    print(f"  -> {path.name}")

    # ── Figure ───────────────────────────────────────────────────────────────
    if not df.empty:
        scs = [s for s in SCENARIOS if s in set(df["scenario"])]
        fig, axes = plt.subplots(2, len(scs), figsize=(4.7 * len(scs), 7.2),
                                 squeeze=False)
        for ci, sc in enumerate(scs):
            sub = df[df["scenario"] == sc]

            ax = axes[0][ci]
            ev = sub[sub["model"] == FUSED]["n_events_test"].dropna()
            if len(ev):
                ax.hist(ev, bins=np.arange(ev.min() - 0.5, ev.max() + 1.5),
                        color=SC_COL.get(sc, "#333"), alpha=0.8,
                        edgecolor="white")
                ax.axvline(ev.median(), color="#c62828", ls="--", lw=1.2,
                           label=f"median {ev.median():.0f}")
                ax.legend(fontsize=7.5)
            ax.set_xlabel("pCR events in the held-out test fold")
            ax.set_ylabel("number of folds")
            ax.set_title(f"{sc} — per-fold test events",
                         color=SC_COL.get(sc, "#333"), fontsize=10)
            ax.grid(axis="y", alpha=0.3)

            ax = axes[1][ci]
            mods = [m for m in ALL_MODELS if m in set(sub["model"])]
            vals = [sub[sub["model"] == m]["epv_realized"].dropna().values
                    for m in mods]
            vals = [v if len(v) else np.array([np.nan]) for v in vals]
            # Tick labels are set separately rather than via boxplot's keyword,
            # which matplotlib renamed from `labels` to `tick_labels` in 3.9.
            bp = ax.boxplot(vals, patch_artist=True,
                            showfliers=False, widths=0.6)
            ax.set_xticks(np.arange(1, len(mods) + 1))
            ax.set_xticklabels(mods)
            for patch, m in zip(bp["boxes"], mods):
                patch.set_facecolor(MOD_COLOR.get(m, "#888"))
                patch.set_alpha(0.75)
            for med in bp["medians"]:
                med.set_color("black")
            ax.axhline(5, color="#c62828", ls="--", lw=1.2,
                       label="EPV = 5 (design cap)")
            ax.set_ylabel("realised events per variable")
            ax.set_title(f"{sc} — realised EPV",
                         color=SC_COL.get(sc, "#333"), fontsize=10)
            ax.tick_params(axis="x", rotation=30)
            ax.legend(fontsize=7.5)
            ax.grid(axis="y", alpha=0.3)
        fig.suptitle(
            "Per-fold pCR event counts and realised events-per-variable",
            fontsize=11, fontweight="bold", y=1.01)
        plt.tight_layout()
        _savefig(fig, fd / "revfig03_epv_per_fold.pdf")

    return df, df_sum


def run_model_comparisons(results, td, fd=None, df_perf=None):
    """
    Paired comparisons of the fused model against every unimodal model on the
    same patients and CV repeats (cluster bootstrap primary, per-repeat
    DeLong secondary).
    """
    import zlib
    frames = []
    for sc, entry in results.items():
        for source, blob in (("consensus", entry["consensus"]),
                             ("discovery", entry["discovery"])):
            if blob is None:
                continue
            mats, y_ref, pid_ref = {}, None, None
            for mod in ALL_MODELS:
                try:
                    rm = _repeat_matrix(blob, source, mod)
                except (KeyError, ValueError):
                    continue
                if rm is None or rm.n_patients == 0:
                    continue
                P, pid, y = rm.P, rm.pids, rm.y
                if pid_ref is None:
                    pid_ref, y_ref = pid, y
                elif not np.array_equal(pid, pid_ref):
                    # Comparisons must be on identical patients. Align rather
                    # than silently comparing different cohorts.
                    common = np.intersect1d(pid_ref, pid)
                    keep_ref = np.isin(pid_ref, common)
                    pid_ref, y_ref = pid_ref[keep_ref], y_ref[keep_ref]
                    for k in mats:
                        mats[k] = mats[k][:, keep_ref]
                    P = P[:, np.isin(pid, common)]
                if mats and P.shape[0] != next(iter(mats.values())).shape[0]:
                    print(f"  [CMP] {sc}/{source}/{mod}: {P.shape[0]} repeats vs "
                          f"{next(iter(mats.values())).shape[0]} — skipped "
                          f"(paired comparison needs identical repeats).")
                    continue
                mats[mod] = P
            if FUSED in mats and len(mats) > 1:
                d = compare_models_repeat(
                    y_ref, mats, reference=FUSED,
                    seed=BOOT_SEED + zlib.crc32(f"cmp|{sc}|{source}".encode()) % 10000)
                if not d.empty:
                    d.insert(0, "source", source)
                    d.insert(0, "scenario", sc)
                    frames.append(d)

    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # ── Fusion-benefit summary: the meta-classifier's added value, plainly ───
    # One row per scenario x source answering the reviewer-facing question
    # directly: what does late fusion buy over (i) the best single modality
    # and (ii) the clinical-only baseline? Both deltas are paired on the
    # same patients; a benefit is claimed only when the interval excludes 0.
    benefit_rows = []
    if not df.empty:
        for (sc, source), sub in df.groupby(["scenario", "source"]):
            sub = sub.reset_index(drop=True)
            if sub["AUROC_comparator"].notna().sum() == 0:
                # Degenerate group (single-class pooled y → DeLong NaN for
                # every comparator): idxmax on all-NaN raises in pandas 3.
                continue
            fused_auroc = float(sub["AUROC_reference"].iloc[0])
            best_i = int(sub["AUROC_comparator"].idxmax())
            best = sub.loc[best_i]
            row = {
                "scenario": sc, "source": source,
                "fused_AUROC": fused_auroc,
                "best_unimodal": best["comparator"],
                "best_unimodal_AUROC": float(best["AUROC_comparator"]),
                "delta_vs_best_unimodal": float(best["delta_AUROC"]),
                "delta_vs_best_CI":
                    f"{best['delta_CI_low']:.3f} to {best['delta_CI_high']:.3f}",
                "delta_vs_best_p_bootstrap": float(best["p_bootstrap"]),
                "delta_vs_best_q_BH": float(best["q_bootstrap_BH"]),
                "verdict_vs_best_unimodal": best["verdict"],
            }
            clin = sub[sub["comparator"] == "Clin"]
            if not clin.empty:
                c = clin.iloc[0]
                row.update({
                    "Clin_AUROC": float(c["AUROC_comparator"]),
                    "delta_vs_Clin": float(c["delta_AUROC"]),
                    "delta_vs_Clin_CI":
                        f"{c['delta_CI_low']:.3f} to {c['delta_CI_high']:.3f}",
                    "delta_vs_Clin_p_bootstrap": float(c["p_bootstrap"]),
                    "delta_vs_Clin_q_BH": float(c["q_bootstrap_BH"]),
                    "verdict_vs_Clin": c["verdict"],
                })
            benefit_rows.append(row)
    df_benefit = pd.DataFrame(benefit_rows)

    wb = openpyxl.Workbook()
    _write_sheet(
        wb, "Model_comparisons", df, first=True,
        note=("PAIRED MODEL COMPARISONS ON THE SAME PATIENTS AND THE SAME "
              "CROSS-VALIDATION REPEATS.\n"
              "AUROC = mean over CV repeats of the pooled out-of-fold AUROC (the "
              "estimand of revision_performance_CI.xlsx). delta = fused minus "
              "comparator.\n"
              f"Primary test: paired patient-level cluster bootstrap ({N_BOOT} "
              "resamples): the same patient resample, carrying all its repeat "
              "predictions, is applied to both models. p_bootstrap is the two-sided "
              "bootstrap p-value; q_bootstrap_BH is Benjamini-Hochberg adjusted "
              "across the comparisons within each scenario and source.\n"
              "Secondary: DeLong's analytic test needs one prediction per patient, so "
              "it is run once per CV repeat and summarised (median p, IQR, fraction "
              "of repeats with p < 0.05). It is descriptive; the verdict never rests "
              "on it.\n"
              "'verdict' is deliberately conservative: any comparison whose 95% "
              "interval includes zero is reported as 'not distinguishable', however "
              "large the point difference. Claims that the integrated model "
              "'outperformed' a comparator should be made only where the interval "
              "excludes zero."))
    _write_sheet(
        wb, "Fusion_benefit", df_benefit,
        note=("THE META-CLASSIFIER'S ADDED VALUE, IN ONE ROW PER SCENARIO.\n"
              "Late fusion (elastic-net over the five calibrated modality "
              "probability streams) is compared against (i) the best single "
              "modality and (ii) the clinical-only baseline, paired on the same "
              "patients and CV repeats (paired cluster bootstrap; BH-adjusted "
              "within scenario x source).\n"
              "A benefit may be claimed ONLY where the paired 95% interval "
              "excludes zero ('verdict'). 'not distinguishable' against the best "
              "single modality with a clear win over the clinical baseline means "
              "the fusion's value is (a) robustness to which modality is best — "
              "not knowable in advance — and (b) the interpretable modality "
              "weights (see the Modality_weight_stability sheet and revfig08), "
              "not raw discrimination.\n"
              "See revfig07 for the visual form of this table."))
    path = td / "revision_model_comparisons.xlsx"
    wb.save(path)
    print(f"  -> {path.name}")

    # ── Figure: AUROC forest + paired differences vs the integrated model ────
    # Left column: pooled OOF AUROC with patient-level bootstrap 95% CI per
    # model (from run_performance_ci). Right column: ΔAUROC (Fused −
    # comparator) with the paired-bootstrap 95% CI; a comparison is coloured
    # only when its interval excludes zero. This is the visual form of the
    # rule that "outperformed" claims require the interval to exclude zero.
    if fd is not None and not df.empty:
        scs = [s for s in SCENARIOS if s in set(df["scenario"])]
        fig, axes = plt.subplots(len(scs), 2,
                                 figsize=(9.2, 2.6 * len(scs)),
                                 squeeze=False)
        for ri, sc in enumerate(scs):
            # Prefer the consensus source; fall back to discovery.
            sub = df[df["scenario"] == sc]
            source = ("consensus" if (sub["source"] == "consensus").any()
                      else "discovery")
            sub = sub[sub["source"] == source]

            # Left: absolute AUROC forest
            axL = axes[ri][0]
            if df_perf is not None and not df_perf.empty:
                perf = df_perf[(df_perf["scenario"] == sc)
                               & (df_perf["source"] == source)]
                perf = (perf.set_index("model")
                        .reindex([m for m in ALL_MODELS
                                  if m in set(perf["model"])])
                        .reset_index())
                ypos = np.arange(len(perf))
                cols = [MOD_COLOR.get(m, "#666") for m in perf["model"]]
                axL.errorbar(perf["AUROC"], ypos,
                             xerr=_asym_err(perf["AUROC"],
                                            perf["AUROC_CI_low"],
                                            perf["AUROC_CI_high"]),
                             fmt="none", ecolor="#444", elinewidth=0.9,
                             capsize=2.5)
                axL.scatter(perf["AUROC"], ypos, s=40, c=cols, zorder=3)
                axL.axvline(0.5, color="#999", lw=0.8, ls=":")
                axL.set_yticks(ypos)
                axL.set_yticklabels(
                    ["Integrated" if m == FUSED else m
                     for m in perf["model"]], fontsize=8)
                axL.invert_yaxis()
                axL.set_xlim(0.35, 1.0)
                axL.set_xlabel("cross-validated AUROC, mean over repeats\n"
                               "(95% patient-level cluster-bootstrap CI)",
                               fontsize=8)
            else:
                axL.axis("off")
            axL.set_title(f"{sc} — performance", color=SC_COL.get(sc, "#333"),
                          fontsize=9.5)
            axL.grid(axis="x", alpha=0.3)

            # Right: paired differences vs the integrated model
            axR = axes[ri][1]
            ypos = np.arange(len(sub))
            colors = []
            for _, r in sub.iterrows():
                if r["verdict"] == "not distinguishable":
                    colors.append("#9a9a9a")
                elif r["delta_AUROC"] > 0:
                    colors.append("#2166ac")   # fused higher
                else:
                    colors.append("#d6604d")   # comparator higher
            axR.errorbar(sub["delta_AUROC"], ypos,
                         xerr=_asym_err(sub["delta_AUROC"],
                                        sub["delta_CI_low"],
                                        sub["delta_CI_high"]),
                         fmt="none", ecolor="#444", elinewidth=0.9,
                         capsize=2.5)
            axR.scatter(sub["delta_AUROC"], ypos, s=40, c=colors, zorder=3)
            for yi, (_, r) in enumerate(sub.iterrows()):
                axR.text(1.02, yi, f"q={r['q_bootstrap_BH']:.3f}",
                         transform=axR.get_yaxis_transform(),
                         fontsize=6.8, va="center", color="#444")
            axR.axvline(0, color="#999", lw=0.8, ls=":")
            axR.set_yticks(ypos)
            axR.set_yticklabels(sub["comparator"], fontsize=8)
            axR.invert_yaxis()
            axR.set_xlabel("ΔAUROC: integrated − comparator "
                           "(paired bootstrap 95% CI)", fontsize=8)
            axR.set_title(f"{sc} — paired comparison ({source})",
                          color=SC_COL.get(sc, "#333"), fontsize=9.5)
            axR.grid(axis="x", alpha=0.3)
        fig.suptitle(
            "Model performance and paired comparisons against the integrated "
            "model\ngrey = 95% interval includes zero (not distinguishable); "
            "q = paired cluster-bootstrap p, BH-adjusted within scenario",
            fontsize=10.5, fontweight="bold", y=1.01)
        plt.tight_layout()
        _savefig(fig, fd / "revfig07_model_comparisons.pdf")

    return df


def run_biomarker_groups(results, td, fd, df_data=None, rules=None,
                         arm_col="Clin_Arm", outcome_col="pCR",
                         s_spec=None, cutpoint_q=0.5, n_perm=1000,
                         adjust_cols=("Clin_ER",)):
    """
    Exploratory biomarker groups (Figure 7) with a treatment-by-group
    interaction test and a cut-point sensitivity analysis.

    Grouping source, in order of preference:
      1. The published S1/S2/S3 decision tree, when the four driver columns are
         present in the dataset (see S_GROUP_SPEC_DEFAULT and --s_group_spec).
      2. Explicit rules supplied via --groups_json.
      3. Tertiles of the model's out-of-fold predicted probability, as a
         fallback that at least exercises the same machinery.

    Note on the cohort: the S-group analysis runs on ALL patients who have the
    four driver variables (the manuscript's Figure 7 uses n=181), not on the
    n=110 complete-case machine-learning cohort. These are different analyses
    on different populations and the output states which was used.
    """
    panel = None
    group_source = None
    rules_applied = pd.DataFrame()
    cutinfo = {}
    sweep = pd.DataFrame()
    perm = {}

    # ── Route 1: the published S1/S2/S3 decision tree ────────────────────────
    if df_data is not None:
        labels, cut = assign_s_groups(df_data, spec=s_spec, q=cutpoint_q)
        if labels is not None:
            d = df_data.copy()
            d["group"] = labels
            panel = pd.DataFrame({
                "patient_id": d["patient_id"].values,
                "pCR": d[outcome_col].values,
                "arm": d[arm_col].values,
                "group": d["group"].values,
            })
            for c in adjust_cols:
                if c in d.columns:
                    panel[c] = d[c].values
            cutinfo = cut
            group_source = (
                f"published S1/S2/S3 decision tree (Figure 7e), shared "
                f"proportional cut-off q={cutpoint_q}")
            rules_applied = pd.DataFrame([
                {"step": 1, "group_assigned": "S1",
                 "rule": (f"{cut['erbb2_scna']['column']} >= "
                          f"{cut['erbb2_scna']['threshold']:.4g} (q="
                          f"{cut['erbb2_scna']['quantile']}) AND "
                          f"{cut['genome_instability']['column']} >= "
                          f"{cut['genome_instability']['threshold']:.4g} (q="
                          f"{cut['genome_instability']['quantile']})"),
                 "description": "high ERBB2 SCNA and high genome instability"},
                {"step": 2, "group_assigned": "S3",
                 "rule": (f"NOT S1 AND {cut['adc_trafficking']['column']} >= "
                          f"{cut['adc_trafficking']['threshold']:.4g} (q="
                          f"{cut['adc_trafficking']['quantile']}) AND "
                          f"{cut['type2_immunity']['column']} <= "
                          f"{cut['type2_immunity']['threshold']:.4g} (q="
                          f"{cut['type2_immunity']['quantile']})"),
                 "description": "high ADC trafficking and low type 2 immunity"},
                {"step": 3, "group_assigned": "S2",
                 "rule": "all remaining classifiable patients",
                 "description": "neither signature pattern"},
                {"step": 4, "group_assigned": "Unclassifiable",
                 "rule": "any driver variable missing",
                 "description": "reported rather than folded into S2, which "
                                "would inflate the S2 denominator"},
            ])
            print(f"  [GROUPS] S1/S2/S3 assigned "
                  f"({dict(pd.Series(labels).value_counts())})")
        else:
            print(f"  [GROUPS] S-group driver columns not found: "
                  f"{cut.get('missing')}")
            print("           Supply the correct column names with "
                  "--s_group_spec, or rules with --groups_json.")

    # ── Route 2: explicit rules ──────────────────────────────────────────────
    if panel is None and rules and df_data is not None:
        labels, rules_applied = assign_groups_from_rules(df_data, rules)
        panel = pd.DataFrame({
            "patient_id": df_data["patient_id"].values,
            "pCR": df_data[outcome_col].values,
            "arm": df_data[arm_col].values,
            "group": labels.values,
        })
        for c in adjust_cols:
            if c in df_data.columns:
                panel[c] = df_data[c].values
        group_source = "pre-specified biomarker rules (--groups_json)"

    # ── Route 3: model-probability tertiles ──────────────────────────────────
    if panel is None:
        entry = results.get("Global")
        if entry is None:
            print("  [GROUPS] Nothing to group on — skipping.")
            return None
        blob = entry["consensus"] or entry["discovery"]
        if blob is None:
            return None
        try:
            if entry["consensus"] is not None:
                pid, y, p, _ = pool_consensus_by_patient(blob, FUSED)
            else:
                pid, y, p, _ = pool_oof_by_patient(blob.get(FUSED, []))
        except (KeyError, ValueError) as e:
            print(f"  [GROUPS] {e}")
            return None
        if len(y) == 0:
            return None
        arm = np.full(len(pid), np.nan)
        # Join arm ONLY when the pooled identifiers are real patient IDs.
        # Old PKLs carry only positional 'test_idx' (positions within the CC
        # cohort); joining those against df_data's patient_id — positions in
        # the 197-row file — would silently misassign arms in the tertile
        # interaction analysis. Leaving arm NaN is tolerated downstream.
        folds0 = (blob.get("folds", []) if entry["consensus"] is not None
                  else blob.get(FUSED, []))
        pids_are_real = bool(folds0) and ("test_pids" in folds0[0])
        if pids_are_real and df_data is not None and arm_col in df_data.columns:
            lut = dict(zip(df_data["patient_id"].astype(int),
                           df_data[arm_col].values))
            arm = np.array([lut.get(int(i), np.nan) for i in pid], dtype=float)
        elif not pids_are_real:
            print("  [GROUPS] fold dicts carry positional test_idx only — "
                  "arm left unassigned rather than misjoined against "
                  "patient_id.")
        panel, cutpoints = risk_tertile_groups(pid, y, p, arm)
        rules_applied = pd.DataFrame([{
            "step": i + 1, "group_assigned": lab, "rule": expr,
            "description": "tertile of out-of-fold predicted probability",
        } for i, (lab, expr) in enumerate([
            ("Low", f"p_pred <= {cutpoints[0]:.4f}"),
            ("Intermediate", f"{cutpoints[0]:.4f} < p_pred <= {cutpoints[1]:.4f}"),
            ("High", f"p_pred > {cutpoints[1]:.4f}")])])
        group_source = ("model-probability tertiles (FALLBACK — the S-group "
                        "driver columns were not found and no --groups_json "
                        "was supplied)")

    panel_clf = panel[panel["group"] != "Unclassifiable"].copy()

    # ── Counts, interaction test, per-group adjusted odds ratios ─────────────
    counts = group_outcome_table(panel_clf, "group", "arm", "pCR")
    inter = interaction_lrt_adjusted(panel_clf, "group", "arm", "pCR",
                                     adjust_cols=adjust_cols)
    per_group = adjusted_arm_or(panel_clf, "group", "arm", "pCR",
                                adjust_cols=adjust_cols)

    # ── Cut-point sensitivity and selective-inference p-value ────────────────
    if cutinfo and "missing" not in cutinfo and df_data is not None:
        print(f"  [GROUPS] cut-point sweep 0.20-0.80 ...")
        sweep, _ = sweep_cutpoints(df_data, arm_col, outcome_col,
                                   spec=s_spec, adjust_cols=adjust_cols)
        if n_perm > 0:
            print(f"  [GROUPS] permutation test over the sweep "
                  f"({n_perm} permutations) ...")
            perm = cutpoint_permutation_p(
                df_data, arm_col, outcome_col, spec=s_spec,
                adjust_cols=adjust_cols, n_perm=n_perm)

    inter_row = pd.DataFrame([{
        "test": "treatment-by-group interaction (likelihood-ratio)",
        "model_null": "pCR ~ group + arm"
                      + ("+ " + " + ".join(inter.get("adjusted_for", []))
                         if inter.get("adjusted_for") else ""),
        "model_full": "null model + group:arm",
        "LRT_statistic": inter["lrt_stat"],
        "df": inter["df"],
        "nominal_p_value": inter["p_value"],
        "n_patients": inter["n"],
        "n_events": inter["n_events"],
        "reference_group": inter.get("reference_group", ""),
        "adjusted_for": ", ".join(inter.get("adjusted_for", [])) or "nothing",
        "cutpoint_quantile_used": cutpoint_q if cutinfo else np.nan,
        "n_cutpoints_searched": perm.get("n_cutpoints_searched", np.nan),
        "best_cutpoint_in_sweep": perm.get("best_cutpoint", np.nan),
        "min_p_across_sweep": perm.get("observed_min_p", np.nan),
        "permutation_corrected_p": perm.get("permutation_p", np.nan),
        "n_permutations": perm.get("n_perm", np.nan),
    }])

    wb = openpyxl.Workbook()
    _write_sheet(
        wb, "Group_rules", rules_applied, first=True,
        note=("EXPLORATORY BIOMARKER GROUPS — HYPOTHESIS-GENERATING ONLY.\n"
              f"Group source: {group_source}.\n"
              "These groups were DERIVED AND EVALUATED IN THE SAME COHORT and are not a "
              "validated decision tool. They require prospective validation before any "
              "clinical use.\n"
              "The rules are printed here with their realised threshold VALUES, exactly "
              "as the code executed them, so they can be audited and applied to an "
              "external cohort."))
    _write_sheet(
        wb, "Group_counts", counts,
        note=("Patient and pCR event counts per group, overall and within each "
              "treatment arm, with Wilson 95% intervals on every rate.\n"
              "Intervals are shown because at these cell sizes a point estimate carries "
              "much less information than it appears to."))
    _write_sheet(
        wb, "Interaction_test", inter_row,
        note=("FORMAL TREATMENT-BY-GROUP INTERACTION TEST.\n"
              "Any differential-benefit interpretation rests on THIS interaction term, "
              "not on the pCR rates being separately elevated within each arm. Two "
              "groups can both respond well to their assigned treatment without there "
              "being evidence that the assignment mattered.\n"
              "IMPORTANT — nominal_p_value is NOT a valid p-value on its own. The "
              "'high'/'low' cut-points were selected by sweeping a proportional cut-off "
              "from 0.20 to 0.80, optimised against this same outcome. "
              "permutation_corrected_p repeats the ENTIRE sweep under permuted "
              "treatment assignment and reports how often a search this wide produces a "
              "result this strong when no interaction exists. That is the value to "
              "quote."))
    _write_sheet(
        wb, "Cutpoint_sensitivity", sweep,
        note=("CUT-POINT SENSITIVITY ANALYSIS.\n"
              "The group construction repeated across the full 0.20-0.80 cut-off grid "
              "described in the Methods. A scheme whose interaction p-value is small at "
              "only one or two of the thirteen cut-points is fragile; one that holds "
              "across most of the grid is a more robust finding. Reporting the whole "
              "sweep lets the reader see which it is, instead of only the selected "
              "value.\n"
              "Group sizes are also reported per cut-point, because a cut-off that "
              "produces a striking odds ratio in a group of six patients is not the same "
              "finding as one that does so in a group of forty."))
    _write_sheet(
        wb, "Per_group_arm_effect", per_group,
        note=("Per-group odds ratio for T-DM1 versus DHP, from a logistic model "
              "adjusted for ER status — matching the manuscript's primary biomarker "
              "model so the numbers are directly comparable with Figure 7f.\n"
              "'sparse_or_separated_cell' flags groups where a group x arm cell has "
              "zero or complete events; there the adjusted model is not estimable and "
              "the reported odds ratio falls back to Haldane-Anscombe-corrected cell "
              "counts, with an interval that should not be read as precise.\n"
              "EXPLORATORY: descriptive contrasts within an exploratory grouping, not "
              "tests of a pre-registered hypothesis."))
    _write_sheet(wb, "Patient_level", panel,
                 note="Per-patient group assignment, arm, adjustment covariates and "
                      "observed pCR, including patients who could not be classified.")
    path = td / "revision_biomarker_groups.xlsx"
    wb.save(path)
    print(f"  -> {path.name}")

    # ── Figure: pCR by group and arm, plus the cut-point sweep ───────────────
    plot = counts[counts["arm"] != "Overall"]
    if not plot.empty:
        groups = sorted(g for g in panel_clf["group"].unique())
        arms = sorted(plot["arm"].unique())
        has_sweep = not sweep.empty
        fig, axes = plt.subplots(
            1, 2 if has_sweep else 1,
            figsize=(11.5 if has_sweep else 6.8, 4.9), squeeze=False)
        ax = axes[0][0]
        width = 0.8 / max(len(arms), 1)
        arm_names = {"0.0": "DHP", "1.0": "T-DM1", "0": "DHP", "1": "T-DM1"}
        for ai, a in enumerate(arms):
            sub = plot[plot["arm"] == a].set_index("group").reindex(groups)
            xs = np.arange(len(groups)) + ai * width - 0.4 + width / 2
            ax.bar(xs, sub["pCR_rate"], width * 0.92,
                   label=arm_names.get(str(a), f"arm {a}"),
                   color=["#2166ac", "#d6604d"][ai % 2], alpha=0.85,
                   edgecolor="white")
            ax.errorbar(xs, sub["pCR_rate"],
                        yerr=_asym_err(sub["pCR_rate"], sub["pCR_ci_low"],
                                       sub["pCR_ci_high"]),
                        fmt="none", ecolor="#333", elinewidth=1.0, capsize=3)
            for x, (_, r) in zip(xs, sub.iterrows()):
                if np.isfinite(r["pCR_rate"]):
                    ax.text(x, 0.02, f"{int(r['n_pCR'])}/{int(r['n'])}",
                            ha="center", va="bottom", fontsize=7, rotation=90,
                            color="white", fontweight="bold")
        ax.set_xticks(np.arange(len(groups)))
        ax.set_xticklabels(groups, fontsize=9)
        ax.set_ylabel("pCR rate (Wilson 95% CI)")
        ax.set_ylim(0, 1.05)
        ax.legend(fontsize=8)
        ax.grid(axis="y", alpha=0.3)
        nom = ("n/a" if not np.isfinite(inter["p_value"])
               else f"{inter['p_value']:.3f}")
        corr = perm.get("permutation_p", np.nan)
        corr_txt = ("" if not np.isfinite(corr)
                    else f"; permutation-corrected p = {corr:.3f}")
        ax.set_title(
            "EXPLORATORY biomarker groups — pCR by group and arm\n"
            f"Interaction LRT nominal p = {nom}{corr_txt}\n"
            f"{inter['n_events']} events in {inter['n']} patients · "
            "derived and evaluated in the same cohort",
            fontsize=9)

        if has_sweep:
            ax2 = axes[0][1]
            ax2.plot(sweep["cutpoint_quantile"], sweep["interaction_p"],
                     "o-", color="#6a1f6a", lw=1.6, ms=5)
            ax2.axhline(0.05, color="#c62828", ls="--", lw=1.1,
                        label="nominal 0.05")
            if np.isfinite(perm.get("best_cutpoint", np.nan)):
                ax2.axvline(perm["best_cutpoint"], color="#2e7d32", ls=":",
                            lw=1.3, label="cut-point minimising p")
            ax2.set_yscale("log")
            ax2.set_xlabel("proportional cut-off quantile")
            ax2.set_ylabel("interaction p-value (log scale)")
            ax2.legend(fontsize=7.5)
            ax2.grid(alpha=0.3)
            ax2.set_title(
                "Cut-point sensitivity\n"
                f"{perm.get('n_cutpoints_searched', 0)} cut-points searched; "
                "the published analysis reports the best of these",
                fontsize=9)
        plt.tight_layout()
        _savefig(fig, fd / "revfig04_biomarker_groups.pdf")

    return {"counts": counts, "interaction": inter, "panel": panel,
            "sweep": sweep, "permutation": perm, "per_group": per_group}


def run_survival(df_data, td, fd, time_col, event_col, variables,
                 group_col=None, adjust_cols=()):
    """Exploratory recurrence analysis with FDR correction and event counts."""
    if df_data is None or time_col not in df_data.columns \
            or event_col not in df_data.columns:
        print(f"  [SURVIVAL] Columns '{time_col}' / '{event_col}' not found "
              "in the dataset — skipping. Pass --survival_time_col and "
              "--survival_event_col to enable this section.")
        return None

    d = df_data.dropna(subset=[time_col, event_col]).copy()
    n_events = int(np.nansum(d[event_col].values))
    adjust_cols = [c for c in adjust_cols if c in d.columns]
    n_cov = 1 + len(adjust_cols)
    print(f"  [SURVIVAL] n={len(d)}, events={n_events}, "
          f"adjusting for {adjust_cols or 'nothing'}")

    res = exploratory_survival(d, time_col, event_col, variables,
                               adjust_cols=adjust_cols)

    header = pd.DataFrame([{
        "n_patients": len(d),
        "n_events": n_events,
        "median_followup": float(np.nanmedian(d[time_col].values)),
        "n_variables_tested": len(res) if res is not None else 0,
        "covariates_per_model": n_cov,
        "events_per_variable": n_events / n_cov if n_cov else np.nan,
        "multiplicity_correction": "Benjamini-Hochberg across tested variables",
        "status": "EXPLORATORY / HYPOTHESIS-GENERATING",
        "power_note": (
            f"{n_events} events across {len(d)} patients. Each model estimates "
            f"{n_cov} coefficients, giving {n_events / n_cov:.1f} events per "
            "variable — far below the conventional minimum of 10. At this ratio "
            "Wald intervals are unreliable, models can approach separation, and "
            "hazard ratios far from 1 reflect instability as much as effect "
            "size. Absence of significance is not evidence of absence of "
            "effect, and nominally significant findings require independent "
            "validation."),
    }])

    wb = openpyxl.Workbook()
    _write_sheet(
        wb, "Survival_overview", header, first=True,
        note=("EXPLORATORY RECURRENCE ANALYSIS.\n"
              "This section is repositioned as exploratory and hypothesis-generating. "
              "Event counts and the realised events-per-variable are stated on every "
              "row, p-values are Benjamini-Hochberg corrected across the variables "
              "tested, and no inferential claim is drawn from this analysis."))
    _write_sheet(
        wb, "Cox_results", res,
        note=("Cox proportional-hazards models (Breslow tie handling), with a log-rank "
              "test additionally reported for binary variables.\n"
              "q_cox_BH and q_logrank_BH are Benjamini-Hochberg adjusted ACROSS the "
              "variables in this table — the multiplicity correction the original "
              "analysis did not apply.\n"
              "'unstable_estimate' flags rows where the hazard ratio exceeds 20 or "
              "falls below 0.05, the standard error is not estimable, or the "
              "events-per-variable is below 5. Such estimates indicate near-separation "
              "rather than a large effect and should not be reported as findings."))
    path = td / "revision_survival_exploratory.xlsx"
    wb.save(path)
    print(f"  -> {path.name}")

    # ── Kaplan-Meier figure ──────────────────────────────────────────────────
    if group_col and group_col in d.columns:
        fig, ax = plt.subplots(figsize=(6.4, 4.8))
        for lvl in sorted(d[group_col].dropna().unique()):
            sub = d[d[group_col] == lvl]
            t, s, _, _ = kaplan_meier(sub[time_col].values, sub[event_col].values)
            if len(t) == 0:
                continue
            ax.step(np.concatenate([[0], t]), np.concatenate([[1.0], s]),
                    where="post", lw=1.8,
                    label=f"{lvl} (n={len(sub)}, "
                          f"{int(np.nansum(sub[event_col].values))} events)")
        lr = logrank_test(d[time_col].values, d[event_col].values,
                          d[group_col].values)
        ax.set_ylim(0, 1.02)
        ax.set_xlabel(time_col)
        ax.set_ylabel("Recurrence-free survival")
        ax.legend(fontsize=8)
        ax.grid(alpha=0.3)
        p_txt = "n/a" if not np.isfinite(lr["p_value"]) else f"{lr['p_value']:.3f}"
        ax.set_title(
            f"EXPLORATORY recurrence-free survival by {group_col}\n"
            f"Log-rank p = {p_txt} · {n_events} events in {len(d)} patients\n"
            "Underpowered; hypothesis-generating only",
            fontsize=9.5)
        plt.tight_layout()
        _savefig(fig, fd / "revfig05_survival_exploratory.pdf")

    return res


# =============================================================================
# SECTION 10 — CLI
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="PREDIX HER2 revision analyses — calibration, patient-level "
                    "bootstrap CIs, selection stability, paired model tests, "
                    "per-fold EPV, exploratory biomarker groups and survival.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    p.add_argument("--results_dir", type=Path, required=True,
                   help="Directory containing {global,dhp,tdm1}/ PKL outputs.")
    p.add_argument("--out_dir", type=Path, default=Path("./report"),
                   help="Report root; tables/revision and figures/revision are "
                        "created inside it.")
    p.add_argument("--data_path", type=Path, default=None,
                   help="Original dataset. Required for the biomarker-group and "
                        "survival sections; the other sections need only the PKLs.")
    p.add_argument("--groups_json", type=Path, default=None,
                   help="JSON file with pre-specified biomarker group rules: a list of "
                        "{\"name\", \"expr\", \"description\"} objects, evaluated in "
                        "order with first match winning. Used only when the S1/S2/S3 "
                        "driver columns are not present.")

    # ── Figure 7 S1/S2/S3 groups ─────────────────────────────────────────────
    p.add_argument("--s_group_spec", type=Path, default=None,
                   help="JSON mapping the four S-group driver roles to column names: "
                        '{"erbb2_scna": "...", "genome_instability": "...", '
                        '"adc_trafficking": "...", "type2_immunity": "..."}. '
                        f"Defaults to {S_GROUP_SPEC_DEFAULT}.")
    p.add_argument("--cutpoint_q", type=float, default=0.5,
                   help="Shared proportional cut-off quantile used for the primary "
                        "S-group assignment. The sensitivity analysis sweeps the full "
                        "0.20-0.80 grid regardless of this value.")
    p.add_argument("--n_perm", type=int, default=1000,
                   help="Permutations for the selective-inference p-value that "
                        "accounts for the cut-point search. Set 0 to skip (fast).")
    p.add_argument("--adjust_cols", nargs="*", default=["Clin_ER"],
                   help="Covariates for the ER-adjusted logistic models, matching the "
                        "manuscript's primary biomarker model.")

    # ── Exploratory survival ─────────────────────────────────────────────────
    p.add_argument("--survival_time_col", default="EFS_months",
                   help="Follow-up time column. The trial endpoint is event-free "
                        "survival (EFS): progression, recurrence, contralateral breast "
                        "cancer, or death from any cause.")
    p.add_argument("--survival_event_col", default="EFS_event",
                   help="Event indicator column (1 = event).")
    p.add_argument("--survival_group_col", default=None,
                   help="Optional column to stratify the Kaplan-Meier plot by.")
    p.add_argument("--survival_adjust_cols", nargs="*",
                   default=["Clin_Arm", "Clin_TUMSIZE", "Clin_ANYNODES", "Clin_ER"],
                   help="Adjustment set for the Cox models, matching the manuscript "
                        "(treatment arm, tumour size, nodal status, ER status). Pass "
                        "no values for univariable models.")

    p.add_argument("--n_boot", type=int, default=N_BOOT,
                   help="Patient-level bootstrap resamples.")
    p.add_argument("--seed", type=int, default=BOOT_SEED,
                   help="Bootstrap seed. Recorded in the outputs for reproducibility.")
    return p.parse_args()


def main():
    global N_BOOT, BOOT_SEED
    args = parse_args()
    N_BOOT = args.n_boot
    BOOT_SEED = args.seed

    td = args.out_dir / "tables" / "revision"
    fd = args.out_dir / "figures" / "revision"
    td.mkdir(parents=True, exist_ok=True)
    fd.mkdir(parents=True, exist_ok=True)

    print("=" * 72)
    print("PREDIX HER2 — REVISION ANALYSES")
    print(f"  Results dir : {args.results_dir}")
    print(f"  Output      : {args.out_dir}")
    print(f"  Bootstrap   : {N_BOOT} patient-level resamples, seed {BOOT_SEED}")
    print("=" * 72)

    results = load_results(args.results_dir)
    print(f"\n[LOAD] scenarios found: {', '.join(results.keys())}")

    df_data = None
    if args.data_path and args.data_path.exists():
        df_data = pd.read_csv(args.data_path, sep="\t")
        # The pipeline assigns patient_id as the row position of the ORIGINAL
        # file, before any filtering, and stores that in every fold's
        # test_pids. Reproduce it exactly here or the group assignments join to
        # the wrong patients.
        if "patient_id" not in df_data.columns:
            df_data["patient_id"] = range(len(df_data))
        # Encode the treatment arm the same way the pipeline does, so the
        # interaction test uses identical coding (DHP = 0, T-DM1 = 1).
        # Tested with is_numeric_dtype rather than `== object`: pandas 3
        # gives string columns a dedicated `str` dtype, so the old check
        # silently skipped the mapping and left 'DHP'/'T-DM1' as text.
        if not pd.api.types.is_numeric_dtype(df_data["Clin_Arm"]):
            df_data["Clin_Arm"] = df_data["Clin_Arm"].map(
                {"DHP": 0, "T-DM1": 1}).astype(float)
        for c in ("Clin_ER",):
            if c in df_data.columns and not pd.api.types.is_numeric_dtype(df_data[c]):
                df_data[c] = df_data[c].map(
                    {"positive": 1, "negative": 0,
                     "Positive": 1, "Negative": 0}).astype(float)
        print(f"[LOAD] dataset: {len(df_data)} patients, "
              f"{int(np.nansum(df_data['pCR'].values))} pCR events")
        # The Figure 7 grouping is defined on every patient with the four
        # driver variables, which in the full trial file is ~181 patients —
        # not the n=110 complete-case machine-learning cohort. Point
        # --data_path at the full 197-patient file to reproduce that analysis.
        n_drv = sum(1 for c in S_GROUP_SPEC_DEFAULT.values()
                    if c in df_data.columns)
        if n_drv == len(S_GROUP_SPEC_DEFAULT):
            classifiable = df_data[list(S_GROUP_SPEC_DEFAULT.values())] \
                .notna().all(axis=1).sum()
            print(f"[LOAD] S-group drivers present; {classifiable} patients "
                  f"classifiable")

    import json
    rules = None
    if args.groups_json and args.groups_json.exists():
        with open(args.groups_json, "r", encoding="utf-8") as f:
            rules = json.load(f)
        print(f"[LOAD] {len(rules)} pre-specified group rules")

    s_spec = None
    if args.s_group_spec and args.s_group_spec.exists():
        with open(args.s_group_spec, "r", encoding="utf-8") as f:
            s_spec = json.load(f)
        print(f"[LOAD] S-group column spec: {s_spec}")

    print("\n[1/6] Patient-level bootstrap confidence intervals ...")
    df_perf = run_performance_ci(results, td)

    print("\n[2/6] Calibration ...")
    run_calibration(results, td, fd)

    print("\n[3/6] Selection and modality-weight stability ...")
    run_stability(results, td, fd)

    print("\n[4/6] Per-fold events and realised EPV ...")
    run_epv(results, td, fd)

    print("\n[5/6] Paired model comparisons ...")
    run_model_comparisons(results, td, fd, df_perf=df_perf)

    print("\n[6/6] Exploratory biomarker groups and interaction test ...")
    run_biomarker_groups(results, td, fd, df_data=df_data, rules=rules,
                         s_spec=s_spec, cutpoint_q=args.cutpoint_q,
                         n_perm=args.n_perm,
                         adjust_cols=tuple(args.adjust_cols))

    if df_data is not None:
        print("\n[extra] Exploratory survival analysis ...")
        candidate_vars = [c for c in df_data.columns
                          if c.startswith(("Clin_", "RNA_", "DNA_",
                                           "Prot_", "WSI_"))
                          and pd.api.types.is_numeric_dtype(df_data[c])]
        run_survival(df_data, td, fd,
                     args.survival_time_col, args.survival_event_col,
                     candidate_vars, group_col=args.survival_group_col,
                     adjust_cols=tuple(args.survival_adjust_cols))

    print("\n" + "=" * 72)
    print("REVISION ANALYSES COMPLETE")
    print(f"  Tables  -> {td}")
    print(f"  Figures -> {fd}")
    print("=" * 72)


if __name__ == "__main__":
    main()
